#!/usr/bin/env python3
"""
QA Regression TestCase Analyzer - Prototype (3 sheets only)
수행: 정량 집계만. LLM 의미 분석 없음.

대상 시트: 1.가입, 2.친구, 3.프로필

컬럼 매핑 (3개 시트 공통):
  col 0  = 분류1
  col 1  = 분류2
  col 2  = 분류3
  col 3  = 분류4 (기능)
  col 4  = 사전 조건
  col 5  = Test Step (체크할 항목)
  col 6  = 기대결과
  col 7  = Android 16 버전 체크
  col 8  = iOS 18 버전 체크
  col 9-26 = 기타 버전 체크 컬럼 (대부분 비어있음)
  col 27 = Priority (P)  → P0~P5 / '-' / 기타
  col 28 = OS Category (C)  → A=Android, I/iOS=iOS, app=공통
  col 29 = Version (V)  → 버전 문자열

TC 판별 규칙 (analyzer.md TC 영역 판별 규칙 반영):
  Step 1 — 통계 영역 경계 감지
    "행 추가시..." 문자열이 분류1(col0)에 나타나는 행을 경계로 삼는다.
    경계 행 및 이후 모든 행은 통계/집계 영역으로 간주해 분석에서 제외한다.

  Step 2 — 경계 이전 행을 개별 판별 (is_tc_row)
    TC 조건: Priority(col27) 정수 0~5  AND  기대결과(col6) 비어있지 않음
    제외 사유별 분류:
      STAT_BOUNDARY : 경계 행 이후 (통계/집계 영역)
      EMPTY_ROW     : 모든 핵심 컬럼이 비어있는 완전 공백 행
      NO_PRIORITY   : Priority 비어있거나 '-'
      INVALID_PRIORITY : Priority가 비정수 또는 0~5 범위 밖
      NO_EXPECTED   : Priority OK, 기대결과 공란 (설명·메모 행)

데이터 시작: 행 인덱스 4 (Excel 5행, 4개 헤더 행 제외)
"""

import difflib
import openpyxl
import csv
import json
import os
import re
from collections import defaultdict

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_FILE = os.path.join(BASE_DIR, 'input', 'Regression TestCase.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

TARGET_SHEETS = [
    '1.가입', '2.친구', '3.프로필',
    '4.채팅목록', '5.일반채팅', '6.비밀채팅',
    '7.팀채팅_채팅방', '8.팀채팅_홈상세', '9.숏폼',
    '10.통화', '11.톡뮤직', '12.이모티콘플러스(+키보드)',
    '13.검색', '14.이모티콘스토어', '15. 알림센터',
    '16.톡클라우드_연락처', '17.지갑', '18.톡캘린더',
    '19.톡클라우드_백복(로직)', '20.톡클라우드_인물분류',
    '21.톡클라우드_버티컬(무료)', '22.톡클라우드_버티컬(유료)',
    '23.죠르디', '24. 더보기', '25. 설정', '26.기타', '27.카나나템플릿',
]

# ── 컬럼 인덱스 상수 ────────────────────────────────────────────────────────────
COL_CAT1     = 0
COL_CAT2     = 1
COL_CAT3     = 2
COL_CAT4     = 3
COL_PRECOND  = 4
COL_STEP     = 5
COL_EXPECTED = 6
COL_PRIORITY = 27
COL_OS       = 28
COL_VERSION  = 29

HEADER_ROWS = 4   # 상위 4행은 헤더, 데이터는 5행부터 (기본값, 동적 탐지로 덮어씀)

# Phase 2 속성 키 목록 (service_summary · attribute_summary · coverage.json 공통)
ATTR_KEYS = [
    'ui_visibility', 'data_change', 'function_behavior',
    'permission_auth', 'exception_error', 'network_server',
    'notification', 'multi_device_os', 'state_persistence', 'content_media',
    'has_precondition', 'automation_candidate', 'manual_required',
]

# 검증 대상 속성만 (복합 속성 조합 대상)
ATTR_KEYS_VERIFICATION = [
    'ui_visibility', 'data_change', 'function_behavior',
    'permission_auth', 'exception_error', 'network_server',
    'notification', 'multi_device_os', 'state_persistence', 'content_media',
]

# 속성 한글 레이블 (compound_attribute 값에 사용)
ATTR_LABEL_KO = {
    'ui_visibility':    'UI노출확인',
    'data_change':      '데이터변경확인',
    'function_behavior':'기능동작확인',
    'permission_auth':  '권한인증확인',
    'exception_error':  '예외에러확인',
    'network_server':   '네트워크서버연동',
    'notification':     '알림푸시확인',
    'multi_device_os':  '멀티디바이스OS',
    'state_persistence':'설정상태유지',
    'content_media':    '콘텐츠미디어확인',
}

# 속성별 키워드 규칙 (step + expected_result 대상)
_ATTR_RULES = {
    'ui_visibility': [
        '노출됨', '노출된다', '노출됩니다', '표시됨', '표시된다', '표시됩니다',
        '보임', '화면 진입', '화면이 표시', '화면으로 이동', '문구가 표시',
        '문구 표시', '아이콘 표시', '아이콘이 표시', '배너가 표시',
        '리스트에 표시', '화면 구성', '화면이 노출',
    ],
    'data_change': [
        '변경됨', '변경된다', '변경됩니다', '저장됨', '저장된다', '저장됩니다',
        '추가됨', '추가된다', '추가됩니다', '삭제됨', '삭제된다', '삭제됩니다',
        '수정됨', '수정된다', '반영됨', '반영된다', '반영됩니다',
        '업데이트됨', '업데이트된다', '생성됨', '생성된다', '해제됨', '해제된다',
    ],
    'function_behavior': [
        '이동됨', '이동된다', '이동됩니다', '실행됨', '실행된다',
        '전송됨', '전송된다', '호출됨', '검색됨', '선택됨',
        '진입됨', '진입된다', '열림', '전환됩니다', '전환된다', '전환됨',
        '랜딩됩니다', '랜딩된다', '랜딩됨',
    ],
    'permission_auth': [
        '권한', '허용', '거부', '로그인', '인증', '약관', '동의', '접근권한',
    ],
    'exception_error': [
        '실패', '오류', '에러', '불가', '제한', '미지원', '차단',
        '팝업이 표시', '팝업이 발생', '초과했습니다', '제한됩니다',
        '초과됩니다', '제한되어', '불가능', '안내',
    ],
    'network_server': [
        '서버', 'api', '네트워크', '동기화', '폴링', '응답', '재시도', '로딩', 'tms',
    ],
    'notification': [
        '푸시', '알림', '배지', '토스트', '팝업', 'n뱃지', '빨간점', '모달',
    ],
    'multi_device_os': [
        'android', 'ios', '기기', '디바이스', '서브디바이스', '멀티 디바이스',
    ],
    'state_persistence': [
        '유지', '재실행', '백그라운드', '복귀', '캐시', '이전 상태', '저장 상태',
    ],
    'content_media': [
        '이미지', '사진', '동영상', '파일', '썸네일', '이모티콘', '미디어', 'gif', '영상',
    ],
}

# 자동화 후보 판단 시그널
_AUTO_CLEAR = [
    '이동됩니다', '이동된다', '표시됩니다', '표시된다',
    '활성화됩니다', '활성화된다', '비활성화됩니다', '비활성화된다',
    '전환됩니다', '전환된다', '추가됩니다', '추가된다',
    '삭제됩니다', '삭제된다', '변경됩니다', '변경된다',
    '완료됩니다', '완료된다', '진행된다', '진행됩니다',
]
_AUTO_VAGUE = ['자연스럽게', '버벅임', '깨지지', '이상 없음', '느낌', '확인 필요']

# 수동 확인 필요 시그널
_MANUAL_SIGNALS = ['시각적', '자연스럽게', '버벅임', '깨지지', '음성 확인', '카메라', '직접 확인']

# ══════════════════════════════════════════════════════════════════════════════
# Phase 4: 사전조건 AI 분석 (규칙 기반)
# ══════════════════════════════════════════════════════════════════════════════
#
# 각 항목:
#   category      : 사전조건 카테고리 키
#   label         : 표시 레이블
#   keywords      : 매칭 키워드 (분석 대상 텍스트 전체에서 검색)
#   trigger_fields: 어떤 컬럼에서 키워드가 발견됐을 때 해당 카테고리를 적용할지
#                   'all' = 모든 컬럼, 'step_exp' = step+expected만
#   infer_desc    : 추론 근거 고정 문구 (키워드 발견 위치를 {field}로 치환)
#
# 우선순위 높은 카테고리가 먼저 매칭됨 (순서 중요)

PRECOND_RULES = [
    # ── 계정 / 인증 ─────────────────────────────────────────────────────────
    {
        'category': '로그인',
        'label': '로그인 완료',
        'keywords': [
            '로그인', '로그인된', '로그인 상태', '로그인 완료',
            '로그인 후', '계정 로그인', '로그인된 상태', '로그아웃',
            '비밀번호', '카카오계정 로그인', '카카오 계정 로그인',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 로그인 또는 계정 관련 동작이 포함되어 있어 로그인 완료 상태가 필요합니다.',
    },
    {
        'category': '특정_계정',
        'label': '특정 계정 필요',
        'keywords': [
            '사전조건의 계정', '탈퇴 계정', '이용제재', '휴면 계정',
            '보호조치 계정', '인증뺏긴', '다른 계정', '특정 계정',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 특정 상태의 계정이 언급되어 있어 해당 계정이 사전에 준비되어야 합니다.',
    },
    {
        'category': '회원가입',
        'label': '회원가입 완료',
        'keywords': [
            '가입 완료', '신규 가입', '최초 가입', '가입된', '기존 가입자',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 가입 여부와 관련된 내용이 포함되어 있어 회원가입이 완료된 상태가 필요합니다.',
    },
    # ── 친구 / 관계 ─────────────────────────────────────────────────────────
    {
        'category': '친구_존재',
        'label': '친구 존재',
        'keywords': [
            '친구a', '친구b', '친구c', '친구d', '친구e', '친구f', '친구h',
            '친구에게', '친구 추가', '친구 목록', '친구탭', '즐겨찾는 친구',
            '차단한 친구', '숨긴 친구', '특정 친구', '친구 존재',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 특정 친구와의 상호작용이 포함되어 있어 해당 친구가 친구 목록에 존재해야 합니다.',
    },
    # ── 채팅방 / 대화방 ────────────────────────────────────────────────────
    {
        'category': '채팅방_존재',
        'label': '채팅방 생성 완료',
        'keywords': [
            '채팅방', '대화방', '1:1 채팅방', '그룹 채팅방', '오픈채팅',
            '팀채팅', '비밀채팅', '채팅방 생성', '기존 채팅방',
            '채팅방 있음', '채팅방 설정', '채팅방 이름',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 채팅방 관련 동작이 포함되어 있어 채팅방이 사전에 생성되어 있어야 합니다.',
    },
    # ── 미디어 / 파일 ─────────────────────────────────────────────────────
    {
        'category': '사진_존재',
        'label': '사진/이미지 존재',
        'keywords': [
            '사진', '이미지', '앨범', '갤러리', '사진 전송', '사진 첨부',
            '프로필 사진', '사진 업로드', '사진 선택', '사진/동영상',
        ],
        'trigger_fields': 'step_exp',
        'infer_desc': '{field}에 사진 관련 동작이 포함되어 있어 사진이 기기에 존재해야 합니다.',
    },
    {
        'category': '동영상_존재',
        'label': '동영상 존재',
        'keywords': [
            '동영상', '비디오', '동영상 전송', '동영상 재생', '동영상 업로드',
            '프로필 동영상', '동영상 첨부',
        ],
        'trigger_fields': 'step_exp',
        'infer_desc': '{field}에 동영상 관련 동작이 포함되어 있어 동영상 파일이 기기에 존재해야 합니다.',
    },
    {
        'category': '파일_존재',
        'label': '파일 존재',
        'keywords': [
            '파일 전송', '파일 첨부', '파일 업로드', '파일 선택',
            '문서 전송', '첨부파일',
        ],
        'trigger_fields': 'step_exp',
        'infer_desc': '{field}에 파일 전송 동작이 포함되어 있어 전송할 파일이 사전에 존재해야 합니다.',
    },
    # ── 권한 / OS ──────────────────────────────────────────────────────────
    {
        'category': '권한_허용',
        'label': '권한 허용',
        'keywords': [
            '권한 허용', '카메라 권한', '연락처 권한', '위치 권한',
            '알림 권한', '마이크 권한', '파일 접근 권한', '권한 설정',
            '접근권한', '권한 부여',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 권한 관련 내용이 포함되어 있어 해당 OS 권한이 허용된 상태여야 합니다.',
    },
    {
        'category': '알림_허용',
        'label': '알림 허용',
        'keywords': [
            '알림 허용', '알림 권한', '푸시 알림', '알림 설정',
            '알림 on', '알림 수신', '알림이 허용',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 알림 수신 관련 내용이 포함되어 있어 알림 권한이 허용된 상태여야 합니다.',
    },
    {
        'category': '특정_OS',
        'label': '특정 OS 필요',
        'keywords': [
            'android only', 'ios only', 'android 한정',
            'os12', 'os13', 'os14', 'ios 18', 'android 16',
            'only android', 'only ios',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 특정 OS 조건이 명시되어 있어 해당 OS 환경이 필요합니다.',
    },
    # ── 설정 / 상태 ────────────────────────────────────────────────────────
    {
        'category': '설정_변경',
        'label': '특정 설정 상태',
        'keywords': [
            '설정 변경 후', '설정된 상태', '설정 완료', '설정 on',
            '설정 off', 'on 상태', 'off 상태', '활성화 상태', '비활성화 상태',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 특정 설정 상태에 대한 언급이 있어 해당 설정이 사전에 변경되어 있어야 합니다.',
    },
    {
        'category': '멀티디바이스',
        'label': '멀티디바이스 환경',
        'keywords': [
            '다른 기기', '기존 기기', '서브디바이스', '멀티 디바이스',
            '다른 단말', '타 기기', '기존에 사용 중인 기기',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 다른 기기와의 연동 내용이 포함되어 있어 복수 기기 환경이 필요합니다.',
    },
    # ── 데이터 / 콘텐츠 ────────────────────────────────────────────────────
    {
        'category': '채팅내역_존재',
        'label': '채팅 내역 존재',
        'keywords': [
            '채팅 내역', '대화 내역', '메시지 내역', '기존 메시지',
            '이전 대화', '채팅 기록',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 기존 채팅 내역 관련 동작이 포함되어 있어 채팅 내역이 사전에 존재해야 합니다.',
    },
    {
        'category': '백업_존재',
        'label': '백업 데이터 존재',
        'keywords': [
            '백업', '백업 데이터', '백업 완료', '백업 존재',
            '클라우드 백업', '톡클라우드', '이전 백업',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 백업 관련 동작이 포함되어 있어 백업 데이터가 사전에 존재해야 합니다.',
    },
    {
        'category': '구독_상태',
        'label': '구독/결제 완료',
        'keywords': [
            '구독', '유료 구독', '톡클라우드 구독', '구독 상태',
            '결제', '인앱 구매', '스토어 구매', '유료 사용자',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 구독 또는 결제 상태에 대한 언급이 있어 사전에 구독/결제가 완료된 상태여야 합니다.',
    },
    {
        'category': '검색결과_존재',
        'label': '검색 결과 존재',
        'keywords': [
            '검색 결과', '검색된', '검색 후', '검색어 입력 후',
            '최근 검색', '검색 완료',
        ],
        'trigger_fields': 'step_exp',
        'infer_desc': '{field}에 검색 결과 관련 동작이 포함되어 있어 검색 결과가 존재하는 상태여야 합니다.',
    },
    # ── 네트워크 ───────────────────────────────────────────────────────────
    {
        'category': '네트워크_연결',
        'label': '네트워크 연결',
        'keywords': [
            '네트워크', '인터넷 연결', 'wi-fi', '와이파이', '모바일 데이터',
            '네트워크 상태', '오프라인', '연결 상태',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 네트워크 관련 조건이 포함되어 있어 특정 네트워크 상태가 필요합니다.',
    },
    # ── 관리자 / 특수 계정 ─────────────────────────────────────────────────
    {
        'category': '관리자_권한',
        'label': '관리자 권한',
        'keywords': [
            '관리자', '어드민', 'admin', '관리자 계정', '관리자 메뉴',
        ],
        'trigger_fields': 'all',
        'infer_desc': '{field}에 관리자 기능 접근이 포함되어 있어 관리자 권한을 가진 계정이 필요합니다.',
    },
    # ── 기타 ──────────────────────────────────────────────────────────────
    {
        'category': '기타',
        'label': '기타 사전조건',
        'keywords': [],   # fallback — 직접 사전조건 컬럼에 값이 있으나 위 카테고리에 미분류
        'trigger_fields': 'precond_only',
        'infer_desc': '사전조건 컬럼에 기재된 내용이 있으나 표준 카테고리로 분류되지 않았습니다.',
    },
]

# 카테고리 키 목록 (순서 유지)
PRECOND_CATEGORIES = [r['category'] for r in PRECOND_RULES]

# 필드 이름 → 한글 표시
_FIELD_LABEL = {
    'precondition': '사전조건 컬럼',
    'test_step':    'Test Step',
    'expected_result': '기대결과',
    'category_1':  '분류1',
    'category_2':  '분류2',
    'category_3':  '분류3',
    'category_4':  '화면전개',
}

# ── Phase 3: 품질 점검 상수 ────────────────────────────────────────────────────

# 이슈 타입 키 (coverage.json 구조와 1:1 대응)
ISSUE_TYPES = [
    'ambiguous_test_step',
    'ambiguous_expected_result',
    'missing_precondition',
    'duplicated_step_expected',
    'multiple_purpose',
    'invalid_priority',
    'invalid_os',
    'invalid_category',
]

# 1. Test Step 모호 표현
#
# 판단 기준:
#   A. 단독 동사형 (완전 일치) — 목적어 없이 동사 하나로만 구성된 경우
#      예) "확인한다" (O)  vs  "국가 리스트를 확인한다" (X, 목적어 있음)
#   B. 구문형 부분 포함 — 자체로 의미가 모호한 표현
#      예) "정상 동작 확인", "정상 동작 여부 확인" 등
#
# 핵심 원칙: 목적어(명사)가 앞에 붙으면 구체적 검증 행위이므로 모호하지 않다.

# A. 단독 동사 — step 전체에서 구두점·번호·공백 제거 후 완전 일치해야 모호로 판정
_VAGUE_STEP_SOLO = [
    '확인한다', '확인', '수행한다', '진입한다', '실행한다',
    '테스트한다', '점검한다', '검증한다',
]

# B. 구문형 — 부분 포함만으로 모호 판정 (목적어 결합과 무관하게 그 자체로 모호)
_VAGUE_STEP_PHRASES = [
    '정상 동작 확인', '정상동작 확인', '정상 여부 확인',
    '정상적으로 동작하는지 확인한다',
]

# 2. 기대결과 모호 표현
_VAGUE_EXPECTED_WORDS = [
    '정상 노출', '정상 동작', '정상적으로 동작', '이상 없음', '이상없음',
    '확인', '정상', '동작함', '노출됨',
]

# 3. 사전조건 없이도 문맥상 필요로 보이는 TC 판단 시그널
_PRECOND_REQUIRED_STEP = [
    '로그인 후', '로그인된 상태', '가입 완료 후', '친구 추가 후',
    '설정 변경 후', '기존 데이터', '이미 존재', '이전에 등록',
    '앞서', '사전에',
]
_PRECOND_REQUIRED_EXPECTED = [
    '이전 상태 유지', '기존 값', '기존 설정', '이전 프로필',
]

# 3-B. 기대결과 내 조건 분기 패턴
# "- 기존 설정이 '...'인 경우\n  ㄴ ..." 형태는 사전 상태에 따라 결과가 달라지는 TC.
# 사전조건 컬럼이 비어있으면 사전 상태가 명시되지 않은 것으로 판단한다.
_PRECOND_IN_EXPECTED = re.compile(
    r'(?:'
    r'기존\s*설정이'          # "기존 설정이 ~ 인 경우"
    r'|인\s*경우\s*[\n\r]'   # "~ 인 경우\n"
    r'|경우\s*[\n\r]\s*ㄴ'   # "경우\n  ㄴ"
    r')',
    re.MULTILINE
)

def _extract_precond_from_text(step: str, expected: str) -> str:
    """
    기대결과 또는 Test Step 텍스트에서 사전조건을 추출한다.

    추출 대상 패턴:
      A. "- 기존 설정이 '...'인 경우" 처럼 조건 분기 줄 → 조건 줄들을 연결
      B. "사전조건:", "전제:" 로 시작하는 줄

    반환값: 추출된 사전조건 문자열 (없으면 빈 문자열)
    """
    for text in [expected, step]:
        if not text:
            continue

        lines = text.splitlines()
        cond_lines = []

        for line in lines:
            stripped = line.strip()
            # 패턴 A: "~ 인 경우" 로 끝나는 줄 (조건 헤더)
            if re.search(r'인\s*경우\s*$', stripped):
                cond_lines.append(stripped)
            # 패턴 B: "사전조건:" / "전제:" 시작
            elif re.match(r'^(사전조건|전제)\s*[:：]', stripped):
                cond_lines.append(re.sub(r'^(사전조건|전제)\s*[:：]\s*', '', stripped))

        if cond_lines:
            return ' / '.join(cond_lines)

    return ''


# 5. 복수 목적 TC 판단: Step에 독립적인 기대결과 구분자가 여러 개
_MULTI_PURPOSE_STEP_MIN_NUMBERED = 4   # "1. ..." 형태 단계가 4개 이상이면 복수 목적 의심

# 7. OS 유효값
_VALID_OS_NORMS = {'android', 'ios', 'common', 'web', 'pc'}


# ── 헬퍼 ───────────────────────────────────────────────────────────────────────

def cell(row, idx):
    """안전하게 row[idx] 를 str로 반환. None 또는 범위 초과 시 빈 문자열."""
    if idx >= len(row) or row[idx] is None:
        return ''
    return str(row[idx]).strip()


# 기대결과 컬럼(col6)에 이 값이 있으면 통계표 행으로 판단
_STAT_KEYWORDS_IN_EXPECTED = {
    'os / browser', 'iteration', '환경 (테스트 수행)', '환경(테스트 수행)',
    'priority', '범위 지정', 'pass rate', 'total', '비율',
    '_x0008_coverage', 'coverage',
}
# 기대결과 컬럼에 단독으로 이 값만 있으면 통계 집계 행
_STAT_SINGLE_CHARS = {'p', 'f', 'b', 'n', 'nr'}


def classify_row(row):
    """
    단일 행의 TC 해당 여부를 판별하고 제외 사유를 반환한다.

    반환값:
      'TC'               — 유효한 테스트 케이스
      'EMPTY_ROW'        — 핵심 컬럼이 모두 비어있는 행
      'NO_PRIORITY'      — Priority 없음 또는 '-'
      'INVALID_PRIORITY' — Priority 비정수 또는 0~5 범위 밖
      'NO_EXPECTED'      — 기대결과 공란 (col3/분류4도 없는 경우)
      'NO_CONTENT'       — Step·분류 모두 없음 (Priority만 있는 행)
      'STAT_ROW'         — 통계표 키워드 행 (경계 이전에도 존재할 수 있음)

    ※ STAT_BOUNDARY 이후 행은 load_sheet()에서 이미 제거됨.
    """
    # ── 1. 완전 공백 행 ────────────────────────────────────────────────────────
    key_cols = [COL_CAT1, COL_STEP, COL_EXPECTED, COL_PRIORITY]
    if all(not cell(row, c) for c in key_cols):
        return 'EMPTY_ROW'

    # ── 2. 기대결과 컬럼 기반 통계표 감지 (경계 마커 이전에도 등장 가능) ─────
    expected_val = cell(row, COL_EXPECTED).lower().strip()
    if expected_val in _STAT_KEYWORDS_IN_EXPECTED:
        return 'STAT_ROW'
    if expected_val in _STAT_SINGLE_CHARS:
        return 'STAT_ROW'

    # ── 3. Priority 판별 ────────────────────────────────────────────────────
    p_str = cell(row, COL_PRIORITY)
    if not p_str or p_str in ('-', 'n/a', 'N/A'):
        return 'NO_PRIORITY'
    try:
        f = float(p_str)
        if not (0.0 <= f <= 5.0 and f == int(f)):
            return 'INVALID_PRIORITY'
    except ValueError:
        return 'INVALID_PRIORITY'

    # ── 4. 기대결과 비어있음 — col3(분류4)에 내용 있으면 TC 인정 ──────────
    if not expected_val:
        if not cell(row, COL_CAT4):
            return 'NO_EXPECTED'

    # ── 5. Step 또는 분류 정보 중 하나 이상 존재해야 TC ─────────────────────
    has_step = bool(cell(row, COL_STEP))
    has_cat  = any(cell(row, c) for c in [COL_CAT1, COL_CAT2, COL_CAT3, COL_CAT4])
    if not has_step and not has_cat:
        return 'NO_CONTENT'

    return 'TC'


def is_tc_row(row):
    """하위 호환용 래퍼. classify_row() 기반."""
    return classify_row(row) == 'TC'


def normalize_priority(raw_str):
    """'1.0' → 'P1', '0' → 'P0', 기타 → 'empty'/'other'."""
    s = raw_str.strip()
    if s in ('', '-', 'n/a', 'N/A'):
        return 'empty'
    try:
        f = float(s)
        if 0.0 <= f <= 5.0 and f == int(f):
            return f'P{int(f)}'
        return 'other'
    except ValueError:
        return 'other'


def normalize_os(raw_str):
    """'A'→'android', 'I'/'iOS'→'ios', 'app'→'common', '-'→'empty', 기타→'other'."""
    s = raw_str.strip()
    if s in ('', '-'):
        return 'empty'
    sl = s.lower()
    if sl in ('a', 'android'):
        return 'android'
    if sl in ('i', 'ios'):
        return 'ios'
    if sl == 'app':
        return 'common'
    if sl == 'web':
        return 'web'
    if sl == 'pc':
        return 'pc'
    return 'other'


def classify_attributes(precond, step, expected):
    """
    precond / step / expected 텍스트를 기반으로 ATTR_KEYS 각 속성의
    True/False 딕셔너리를 반환한다.
    """
    text = ' '.join([precond, step, expected]).lower()

    result = {}

    # A. 검증 대상 속성 (키워드 매칭)
    for attr, keywords in _ATTR_RULES.items():
        result[attr] = any(kw.lower() in text for kw in keywords)

    # B-1. 사전조건 있음: precond 컬럼에 의미 있는 내용
    result['has_precondition'] = bool(precond.strip())

    # B-2. 자동화 후보: 기대결과가 명확한 키워드 포함 + 모호 표현 없음
    has_clear = any(kw in expected for kw in _AUTO_CLEAR)
    has_vague = any(kw in text for kw in _AUTO_VAGUE)
    result['automation_candidate'] = has_clear and not has_vague

    # B-3. 수동 확인 필요: 수동 시그널 포함, 또는 이미지/동영상 포함 & 기대결과 모호
    result['manual_required'] = any(kw in text for kw in _MANUAL_SIGNALS)

    return result


def infer_preconditions(tc_dict):
    """
    Phase 4: 단일 TC 딕셔너리에서 사전조건 카테고리를 추론한다.

    분석 대상: precondition, category_1~4, test_step, expected_result

    반환값: list of dict
      {
        'category'   : str,   # 카테고리 키
        'label'      : str,   # 표시 레이블
        'evidence'   : str,   # 근거 문장 (키워드가 발견된 원문)
        'source_field': str,  # 키워드가 발견된 컬럼명
        'keyword'    : str,   # 매칭된 키워드
        'reason'     : str,   # 추론 근거 문구
        'confidence' : str,   # 'high' | 'medium'
      }
    """
    precond  = tc_dict.get('precondition', '') or ''
    step     = tc_dict.get('test_step', '') or ''
    expected = tc_dict.get('expected_result', '') or ''
    cat1     = tc_dict.get('category_1', '') or ''
    cat2     = tc_dict.get('category_2', '') or ''
    cat3     = tc_dict.get('category_3', '') or ''
    cat4     = tc_dict.get('category_4', '') or ''

    # 필드별 텍스트 딕셔너리
    fields_all = {
        'precondition':    precond,
        'test_step':       step,
        'expected_result': expected,
        'category_1':      cat1,
        'category_2':      cat2,
        'category_3':      cat3,
        'category_4':      cat4,
    }
    fields_step_exp = {
        'test_step':       step,
        'expected_result': expected,
    }
    fields_precond_only = {
        'precondition': precond,
    }

    results  = []
    seen_cat = set()   # 카테고리 중복 방지

    for rule in PRECOND_RULES:
        cat      = rule['category']
        keywords = rule['keywords']
        trigger  = rule['trigger_fields']

        if cat in seen_cat:
            continue

        # trigger_fields 에 따라 탐색 범위 결정
        if trigger == 'all':
            search_fields = fields_all
        elif trigger == 'step_exp':
            search_fields = fields_step_exp
        elif trigger == 'precond_only':
            search_fields = fields_precond_only
        else:
            search_fields = fields_all

        # fallback 카테고리(기타): 키워드 없음 — precond에 값이 있고 기타 카테고리 미분류 시
        if cat == '기타':
            if precond.strip() and not seen_cat - {'기타'}:
                results.append({
                    'category':    cat,
                    'label':       rule['label'],
                    'evidence':    precond.strip()[:100],
                    'source_field': 'precondition',
                    'keyword':     '',
                    'reason':      rule['infer_desc'].replace('{field}', '사전조건 컬럼'),
                    'confidence':  'medium',
                })
                seen_cat.add(cat)
            continue

        # 키워드 매칭
        for kw in keywords:
            kw_lower = kw.lower()
            for fname, ftext in search_fields.items():
                if kw_lower in ftext.lower():
                    # 근거 문장: 키워드가 포함된 문장 추출
                    evidence = _extract_evidence(ftext, kw)
                    field_label = _FIELD_LABEL.get(fname, fname)
                    # 사전조건 컬럼에서 발견 → confidence high
                    conf = 'high' if fname == 'precondition' else 'medium'
                    results.append({
                        'category':    cat,
                        'label':       rule['label'],
                        'evidence':    evidence[:120],
                        'source_field': fname,
                        'keyword':     kw,
                        'reason':      rule['infer_desc'].replace('{field}', field_label),
                        'confidence':  conf,
                    })
                    seen_cat.add(cat)
                    break
            if cat in seen_cat:
                break

    return results


def _extract_evidence(text: str, keyword: str) -> str:
    """키워드를 포함하는 문장(줄)을 추출한다. 없으면 keyword 자체를 반환."""
    kw_lower = keyword.lower()
    for line in text.splitlines():
        if kw_lower in line.lower():
            return line.strip()
    return keyword


def detect_quality_issues(tc):
    """
    Phase 3: 단일 TC 딕셔너리를 받아 품질 이슈 목록을 반환한다.
    반환값: list of (issue_type: str, issue_reason: str)
    tc 키: service_name, row_number, category_1~4, precondition,
            test_step, expected_result, priority, os
    """
    issues = []
    step     = tc['test_step']
    expected = tc['expected_result']
    precond  = tc['precondition']
    cat1     = tc['category_1']
    cat2     = tc['category_2']
    priority = tc['priority']
    os_norm  = tc['os']

    step_l     = step.strip().lower()
    expected_l = expected.strip().lower()

    # ── 1. Test Step 모호 ────────────────────────────────────────────────────
    if not step.strip():
        issues.append(('ambiguous_test_step', 'Test Step이 비어 있음'))
    else:
        # A. 단독 동사 완전 일치: 번호·구두점·공백을 제거한 뒤 비교
        #    "1. 확인한다." → "확인한다"  /  "국가 리스트를 확인한다" → 제거 후 길어서 불일치
        step_stripped = re.sub(r'^\s*\d+\s*[.)]\s*', '', step.strip())  # 앞 번호 제거
        step_core = re.sub(r'[\s.,。]+$', '', step_stripped).strip()    # 뒤 구두점 제거
        solo_hit = step_core.lower() in _VAGUE_STEP_SOLO

        # B. 구문형 부분 포함
        phrase_hit = any(p in step_l for p in _VAGUE_STEP_PHRASES)

        if solo_hit:
            issues.append(('ambiguous_test_step',
                            f'목적어 없이 동사만으로 구성된 Test Step: "{step_core}"'))
        elif phrase_hit:
            issues.append(('ambiguous_test_step',
                            f'모호한 Test Step 표현 포함: "{step[:60]}"'))

    # ── 2. 기대결과 모호 ────────────────────────────────────────────────────
    if not expected.strip():
        issues.append(('ambiguous_expected_result', '기대결과가 비어 있음'))
    elif any(w in expected_l for w in _VAGUE_EXPECTED_WORDS):
        # "정상 노출됨" 처럼 단독으로만 있을 때만 모호로 판단
        # (다른 구체적 설명이 함께 있으면 제외)
        cleaned = re.sub(r'[.\s]', '', expected_l)
        vague_only = any(
            re.sub(r'[.\s]', '', w) == cleaned
            for w in _VAGUE_EXPECTED_WORDS
        )
        if vague_only:
            issues.append(('ambiguous_expected_result',
                            f'기대결과가 모호한 단어만으로 구성: "{expected[:60]}"'))

    # ── 3. 사전조건 없음 (문맥상 필요) ─────────────────────────────────────
    if not precond.strip():
        # 3-B 우선: 기대결과 내 조건 분기 패턴
        #   "기존 설정이 '...'인 경우\n  ㄴ ..." 처럼
        #   사전 상태에 따라 결과가 달라지는 구조 → 기대결과 안에 사전조건이 내재됨
        has_cond_in_expected = bool(_PRECOND_IN_EXPECTED.search(expected))

        if has_cond_in_expected:
            cond_line = next(
                (l.strip() for l in expected.splitlines() if '인 경우' in l),
                ''
            )
            issues.append(('missing_precondition',
                            f'기대결과 내 조건 분기 포함 — 사전조건 컬럼 분리 필요'
                            f' (예: "{cond_line[:60]}")'))
        else:
            # 3-A. Step/기대결과 키워드 기반 시그널
            needs_keyword = (
                any(w in step_l for w in _PRECOND_REQUIRED_STEP) or
                any(w in expected_l for w in _PRECOND_REQUIRED_EXPECTED)
            )
            if needs_keyword:
                issues.append(('missing_precondition',
                                'Step/기대결과에 사전 상태 언급이 있으나 사전조건 비어 있음'))

    # ── 4. Step ≈ Expected (중복) ─────────────────────────────────────────
    if step.strip() and expected.strip():
        ratio = difflib.SequenceMatcher(
            None,
            re.sub(r'\s+', '', step_l),
            re.sub(r'\s+', '', expected_l)
        ).ratio()
        if ratio >= 0.85:
            issues.append(('duplicated_step_expected',
                            f'Test Step과 기대결과 유사도 {ratio:.0%}'))

    # ── 5. 복수 목적 TC ─────────────────────────────────────────────────────
    # 판단 기준:
    #   A. compound_attribute 있음 → 목적이 속성으로 명확히 분류됨 → 이슈 아님
    #   B. 경계값 테스트 패턴 → 같은 대상을 다른 값으로 순차 검증 → 이슈 아님
    #      (기대결과의 모든 줄에 공통 핵심 단어가 있거나,
    #       기대결과가 동일 상태의 반복 패턴 - 활성화/비활성화 교차 등)
    #   C. 위 두 조건 모두 해당 없음 → 목적 불명확 혼재 → 이슈
    numbered_steps    = re.findall(r'(?:^|\n)\s*\d+\s*[.)]\s+\S', step)
    numbered_expected = re.findall(r'(?:^|\n)\s*\d+\s*[.)]\s+\S', expected)
    compound_attr = tc.get('compound_attribute', '').strip()

    if (len(numbered_steps) >= _MULTI_PURPOSE_STEP_MIN_NUMBERED and
            len(numbered_expected) >= _MULTI_PURPOSE_STEP_MIN_NUMBERED):

        if compound_attr:
            pass  # A. 목적 명확 → 이슈 없음
        else:
            # B. 경계값 테스트 패턴 감지
            #    step 줄들의 과반수 이상에서 같은 의미 단어(2자↑)가 2개 이상 공통
            #    → 같은 대상을 다른 값으로 반복 검증하는 단일 목적 테스트
            step_lines = [l.strip() for l in re.split(r'\n\s*\d+\s*[.)]', step) if l.strip()]
            if len(step_lines) >= 2:
                word_sets = [set(re.findall(r'[가-힣a-zA-Z]{2,}', l)) for l in step_lines]
                # 과반수 이상 줄에 등장하는 단어 집합
                from collections import Counter as _C
                word_freq = _C(w for ws in word_sets for w in ws)
                majority_threshold = max(2, len(step_lines) // 2)
                majority_words = {w for w, cnt in word_freq.items() if cnt >= majority_threshold}
                is_boundary_test = len(majority_words) >= 2
            else:
                is_boundary_test = False

            if not is_boundary_test:
                issues.append(('multiple_purpose',
                                f'Step {len(numbered_steps)}단계 / 기대결과 {len(numbered_expected)}단계: '
                                '복합 속성으로 분류 불가 — 목적 불명확'))

    # ── 6. Priority 이상값 ───────────────────────────────────────────────────
    # is_tc_row()를 통과한 행이므로 여기서는 'empty'/'other' 만 점검
    if priority in ('empty', 'other'):
        issues.append(('invalid_priority',
                        f'Priority 값이 유효하지 않음: "{tc.get("priority_raw","")}"'))

    # ── 7. OS 이상값 ─────────────────────────────────────────────────────────
    if os_norm not in _VALID_OS_NORMS:
        issues.append(('invalid_os',
                        f'OS 값이 유효하지 않음: "{tc.get("os_raw","")}"'))

    # ── 8. 분류 비어 있음 ────────────────────────────────────────────────────
    if not cat1.strip() or not cat2.strip():
        issues.append(('invalid_category',
                        f'분류1="{cat1}" / 분류2="{cat2}" — 주요 분류가 비어 있음'))

    return issues


def forward_fill(rows, col_indices):
    """지정 컬럼들에 대해 위→아래 방향으로 빈 셀을 직전 비어있지 않은 값으로 채운다."""
    last = {c: None for c in col_indices}
    filled = []
    for row in rows:
        row = list(row)
        # 행 길이 보장
        while len(row) <= max(col_indices):
            row.append(None)
        for c in col_indices:
            v = row[c]
            if v is not None and str(v).strip() != '':
                last[c] = str(v).strip()
                # 원본값 유지 (str로 통일)
                row[c] = last[c]
            else:
                row[c] = last[c]   # None 이면 None 그대로 (아직 채울 값 없음)
        filled.append(row)
    return filled


def load_sheet(wb, sheet_name):
    """
    시트를 읽어 TC 분석 대상 행만 반환한다.

    처리 순서:
      1. 헤더 행 동적 탐지 ('분류1' 셀이 있는 행)
      2. 통계 영역 경계 감지:
         - '행 추가시...' 마커 우선 사용
         - 마커 없으면 통계 키워드 행 연속 등장 시점으로 판단
      3. Forward Fill 적용 (분류 컬럼 계단식 복원)

    반환값: (tc_candidate_rows, exclusion_stats, first_excel_row, last_excel_row)
    """
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # ── Step 1: 헤더 행 동적 탐지 ─────────────────────────────────────────────
    header_idx = HEADER_ROWS - 1  # 기본값 (0-based)
    for i, row in enumerate(all_rows[:10]):
        vals = [str(v or '').strip() for v in row[:5]]
        if '분류1' in vals:
            header_idx = i
            break

    data_rows = all_rows[header_idx + 1:]
    total_data = len(data_rows)

    # ── Step 2: 통계 영역 경계 감지 ───────────────────────────────────────────
    boundary_idx = None

    # 2-a: '행 추가시' 마커
    for i, row in enumerate(data_rows):
        if '행 추가시' in str(row[COL_CAT1] or '').strip():
            boundary_idx = i
            break

    # 2-b: 마커 없으면 통계 키워드 행이 첫 등장하는 시점
    if boundary_idx is None:
        for i, row in enumerate(data_rows):
            exp_l = str(row[COL_EXPECTED] or '').strip().lower()
            if exp_l in _STAT_SINGLE_CHARS:
                boundary_idx = i
                break

    stat_boundary_count = 0
    if boundary_idx is not None:
        stat_boundary_count = total_data - boundary_idx
        data_rows = data_rows[:boundary_idx]

    # 데이터 행 범위 (Excel 행 번호, 1-based)
    first_excel_row = header_idx + 2   # 헤더 다음 줄
    last_excel_row  = first_excel_row
    for i in range(len(data_rows) - 1, -1, -1):
        if any(c for c in data_rows[i] if c is not None and str(c).strip()):
            last_excel_row = header_idx + 1 + i + 1
            break

    # ── Step 3: Forward Fill (분류 컬럼 계단식 복원) ─────────────────────────
    data_rows = forward_fill(data_rows, [COL_CAT1, COL_CAT2, COL_CAT3, COL_CAT4])

    # ── Step 4: 개별 행 제외 사유 집계 ────────────────────────────────────────
    exclusion_stats = {
        'STAT_BOUNDARY'    : stat_boundary_count,
        'STAT_ROW'         : 0,
        'EMPTY_ROW'        : 0,
        'NO_PRIORITY'      : 0,
        'INVALID_PRIORITY' : 0,
        'NO_EXPECTED'      : 0,
        'NO_CONTENT'       : 0,
    }
    for row in data_rows:
        reason = classify_row(row)
        if reason != 'TC':
            exclusion_stats[reason] += 1

    return data_rows, exclusion_stats, first_excel_row, last_excel_row


# ── 메인 분석 루프 ──────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Input  : {INPUT_FILE}")
    print(f"Output : {OUTPUT_DIR}")
    print()

    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

    # 결과 저장용
    tc_master_rows = []
    service_stats  = {}
    category_counts = defaultdict(int)
    # 전체 제외 통계 (시트 합산)
    total_read       = 0
    total_exclusions = defaultdict(int)

    # ── 시트별 처리 ────────────────────────────────────────────────────────────
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] 시트 '{sheet_name}' 없음, 건너뜀.")
            continue

        print(f"처리 중: {sheet_name}")
        data_rows, excl, first_row, last_row = load_sheet(wb, sheet_name)

        # 이 시트에서 읽은 전체 원본 데이터 행 수 = TC후보 + 통계경계 이후
        sheet_read = len(data_rows) + excl['STAT_BOUNDARY']
        total_read += sheet_read
        for k, v in excl.items():
            total_exclusions[k] += v
        sheet_tc = sum(1 for r in data_rows if is_tc_row(r))
        print(f"  행범위: {first_row}~{last_row}  |  원본 데이터 행: {sheet_read}  |  TC 판정: {sheet_tc}  |  제외: {sheet_read - sheet_tc}")

        stats = {
            'service_name': sheet_name,
            'total_tc': 0,
            'first_row': first_row,
            'last_row': last_row,
            'priority': defaultdict(int),
            'os': defaultdict(int),
            'attributes': defaultdict(int),           # Phase 2 단일 속성
            'compound_attributes': defaultdict(int),  # Phase 2 복합 속성 조합
            'precond_categories': defaultdict(int),   # Phase 4 사전조건 카테고리 분포
        }

        # Excel 행 번호는 헤더 4행 + 1-based
        for rel_idx, row in enumerate(data_rows):
            excel_row = HEADER_ROWS + rel_idx + 1  # 1-based Excel 행 번호

            if not is_tc_row(row):
                continue

            # ── 필드 추출 ──────────────────────────────────────────────────────
            cat1    = cell(row, COL_CAT1) or ''
            cat2    = cell(row, COL_CAT2) or ''
            cat3    = cell(row, COL_CAT3) or ''
            cat4    = cell(row, COL_CAT4) or ''
            precond = cell(row, COL_PRECOND)
            step    = cell(row, COL_STEP)
            expected = cell(row, COL_EXPECTED)
            p_raw   = cell(row, COL_PRIORITY)
            os_raw  = cell(row, COL_OS)
            version = cell(row, COL_VERSION)

            p_norm  = normalize_priority(p_raw)
            os_norm = normalize_os(os_raw)

            # 사전조건 자동 추출 (precond 비어있을 때 step/expected에서 추출)
            resolved_precond = precond if precond.strip() else (
                _extract_precond_from_text(step, expected) or precond
            )

            # Phase 2: 속성 분류 (추출된 사전조건 기준)
            attrs = classify_attributes(resolved_precond, step, expected)

            # 복합 속성 계산: multiple_purpose 여부와 무관하게
            # 검증 대상 속성 중 활성화된 것이 2개 이상이면 복합 속성 조합을 기록
            active_verif = [k for k in ATTR_KEYS_VERIFICATION if attrs.get(k)]
            compound_attr = (
                '+'.join(ATTR_LABEL_KO[k] for k in active_verif)
                if len(active_verif) >= 2 else ''
            )

            # ── 집계 ───────────────────────────────────────────────────────────
            stats['total_tc'] += 1
            stats['priority'][p_norm] += 1
            stats['os'][os_norm] += 1
            for k, v in attrs.items():
                if v:
                    stats['attributes'][k] += 1
            if compound_attr:
                stats['compound_attributes'][compound_attr] += 1

            # Phase 4: 사전조건 추론
            temp_tc = {
                'precondition': resolved_precond,
                'test_step': step, 'expected_result': expected,
                'category_1': cat1, 'category_2': cat2,
                'category_3': cat3, 'category_4': cat4,
            }
            precond_inferences = infer_preconditions(temp_tc)
            # 카테고리 분포 집계 (TC당 1회 카운트)
            for inf in precond_inferences:
                stats['precond_categories'][inf['category']] += 1
            # 추론된 카테고리 목록 (쉼표 구분)
            inferred_cats = ','.join(dict.fromkeys(i['category'] for i in precond_inferences))

            cat_key = (sheet_name, cat1, cat2, cat3, cat4)
            category_counts[cat_key] += 1

            row_dict = {
                'service_name'    : sheet_name,
                'row_number'      : excel_row,
                'category_1'      : cat1,
                'category_2'      : cat2,
                'category_3'      : cat3,
                'category_4'      : cat4,
                'precondition'    : resolved_precond,
                'test_step'       : step,
                'expected_result' : expected,
                'priority'        : p_norm,
                'priority_raw'    : p_raw,
                'os'              : os_norm,
                'os_raw'          : os_raw,
                'version'         : version,
                'compound_attribute'  : compound_attr,
                'precond_categories'  : inferred_cats,  # Phase 4
            }
            # Phase 2: tc_master에 속성 플래그 추가 (1/0)
            for k in ATTR_KEYS:
                row_dict[k] = 1 if attrs.get(k) else 0
            tc_master_rows.append(row_dict)

        service_stats[sheet_name] = stats
        print(f"  → TC {stats['total_tc']}건  "
              f"P분포: {dict(stats['priority'])}  "
              f"OS분포: {dict(stats['os'])}")

    # ── 출력 파일 1: tc_master.csv ─────────────────────────────────────────────
    tc_master_path = os.path.join(OUTPUT_DIR, 'tc_master.csv')
    tc_fields = [
        'service_name', 'row_number',
        'category_1', 'category_2', 'category_3', 'category_4',
        'precondition', 'test_step', 'expected_result',
        'priority', 'priority_raw', 'os', 'os_raw', 'version',
        'compound_attribute',   # Phase 2: 복합 속성 조합
        'precond_categories',   # Phase 4: 추론된 사전조건 카테고리 (쉼표 구분)
    ] + ATTR_KEYS  # Phase 2: 속성 플래그 컬럼
    with open(tc_master_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=tc_fields)
        writer.writeheader()
        writer.writerows(tc_master_rows)
    print(f"\n[OK] tc_master.csv  → {len(tc_master_rows)}행")

    # ── 출력 파일 2: service_summary.csv ──────────────────────────────────────
    svc_path = os.path.join(OUTPUT_DIR, 'service_summary.csv')
    svc_fields = (
        ['service_name', 'total_tc',
         'p0_count', 'p1_count', 'p2_count', 'p3_count',
         'p4_count', 'p5_count',
         'priority_empty_count', 'priority_other_count',
         'android_count', 'ios_count', 'common_count',
         'web_count', 'pc_count',
         'os_empty_count', 'os_other_count']
        + [f'{k}_count' for k in ATTR_KEYS]  # Phase 2
    )
    with open(svc_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=svc_fields)
        writer.writeheader()
        for sname in TARGET_SHEETS:
            if sname not in service_stats:
                continue
            s = service_stats[sname]
            p = s['priority']
            o = s['os']
            a = s['attributes']
            row = {
                'service_name'         : sname,
                'total_tc'             : s['total_tc'],
                'p0_count'             : p.get('P0', 0),
                'p1_count'             : p.get('P1', 0),
                'p2_count'             : p.get('P2', 0),
                'p3_count'             : p.get('P3', 0),
                'p4_count'             : p.get('P4', 0),
                'p5_count'             : p.get('P5', 0),
                'priority_empty_count' : p.get('empty', 0),
                'priority_other_count' : p.get('other', 0),
                'android_count'        : o.get('android', 0),
                'ios_count'            : o.get('ios', 0),
                'common_count'         : o.get('common', 0),
                'web_count'            : o.get('web', 0),
                'pc_count'             : o.get('pc', 0),
                'os_empty_count'       : o.get('empty', 0),
                'os_other_count'       : o.get('other', 0),
            }
            for k in ATTR_KEYS:
                row[f'{k}_count'] = a.get(k, 0)
            writer.writerow(row)
    print(f"[OK] service_summary.csv  → {len(service_stats)}행")

    # ── 출력 파일 3: category_summary.csv ─────────────────────────────────────
    cat_path = os.path.join(OUTPUT_DIR, 'category_summary.csv')
    with open(cat_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=[
            'service_name', 'category_1', 'category_2',
            'category_3', 'category_4', 'tc_count',
        ])
        writer.writeheader()
        for (sname, c1, c2, c3, c4), cnt in sorted(category_counts.items()):
            writer.writerow({
                'service_name': sname,
                'category_1'  : c1,
                'category_2'  : c2,
                'category_3'  : c3,
                'category_4'  : c4,
                'tc_count'    : cnt,
            })
    print(f"[OK] category_summary.csv  → {len(category_counts)}행")

    # ── 출력 파일 4: coverage.json ─────────────────────────────────────────────
    total_tc = sum(s['total_tc'] for s in service_stats.values())
    coverage = {
        'total_tc' : total_tc,
        'services' : [],
    }
    for sname in TARGET_SHEETS:
        if sname not in service_stats:
            continue
        s = service_stats[sname]
        p = s['priority']
        o = s['os']
        a  = s['attributes']
        ca = s['compound_attributes']
        coverage['services'].append({
            'service_name': sname,
            'total_tc'    : s['total_tc'],
            'first_row'   : s.get('first_row', 5),
            'last_row'    : s.get('last_row', 5),
            'priority': {
                'P0': p.get('P0', 0), 'P1': p.get('P1', 0),
                'P2': p.get('P2', 0), 'P3': p.get('P3', 0),
                'P4': p.get('P4', 0), 'P5': p.get('P5', 0),
                'empty': p.get('empty', 0), 'other': p.get('other', 0),
            },
            'os': {
                'android': o.get('android', 0), 'ios': o.get('ios', 0),
                'common' : o.get('common',  0), 'web': o.get('web',  0),
                'pc'     : o.get('pc',      0), 'empty': o.get('empty', 0),
                'other'  : o.get('other',   0),
            },
            'attributes': {k: a.get(k, 0) for k in ATTR_KEYS},  # Phase 2 단일
            'compound_attributes': dict(sorted(                   # Phase 2 복합
                ca.items(), key=lambda x: -x[1]
            )),
            'precond_distribution': dict(sorted(                  # Phase 4
                s['precond_categories'].items(), key=lambda x: -x[1]
            )),
        })

    cov_path = os.path.join(OUTPUT_DIR, 'coverage.json')
    with open(cov_path, 'w', encoding='utf-8') as f:
        json.dump(coverage, f, ensure_ascii=False, indent=2)
    print(f"[OK] coverage.json")

    # ── 출력 파일 5: attribute_summary.csv (Phase 2) ──────────────────────────
    _write_attribute_summary(service_stats)

    # ── 출력 파일 5b: precondition_summary.csv (Phase 4) ──────────────────────
    _write_precondition_summary(service_stats, tc_master_rows)

    # ── 출력 파일 5c: tc_reconstruction.csv (Phase 4-B) ────────────────────────
    _write_reconstruction_summary(tc_master_rows)

    # ── 출력 파일 5d: tc_context_linking.csv (Phase 4-C) ───────────────────────
    _write_context_linking(tc_master_rows)

    # ── 출력 파일 5e: ai_standard_tc.csv (공통 데이터 레이어) ──────────────────
    _write_ai_standard_tc(tc_master_rows)

    # ── Phase 3: 품질 점검 ─────────────────────────────────────────────────────
    quality_issues, quality_stats = _run_quality_check(tc_master_rows)

    # ── 출력 파일 6 (Phase 3): quality_issues.csv ──────────────────────────────
    _write_quality_issues(quality_issues)

    # coverage.json에 quality 블록 추가 후 재저장
    _patch_coverage_quality(cov_path, quality_stats, service_stats)

    # 제외 통계 dict 구성 → summary.md 전달용
    excl_report = dict(total_exclusions)
    excl_report['total_read'] = total_read
    excl_report['total_tc']   = total_tc

    # ── 출력 파일 7: summary.md ────────────────────────────────────────────────
    _write_summary(service_stats, total_tc, quality_stats, excl_report)

    total_excluded = total_read - total_tc
    print(f"\n{'='*55}")
    print(f"분석 완료")
    print(f"  전체 읽은 행 수  : {total_read}")
    print(f"  TC 판정 행 수    : {total_tc}")
    print(f"  제외 행 수       : {total_excluded}")
    print(f"  ┌ 통계 영역 경계 이후 : {total_exclusions['STAT_BOUNDARY']}")
    print(f"  ├ 통계표 키워드 행    : {total_exclusions['STAT_ROW']}")
    print(f"  ├ Priority 없음       : {total_exclusions['NO_PRIORITY']}")
    print(f"  ├ Priority 비허용값   : {total_exclusions['INVALID_PRIORITY']}")
    print(f"  ├ 기대결과 공란       : {total_exclusions['NO_EXPECTED']}")
    print(f"  ├ Step·분류 모두 없음 : {total_exclusions['NO_CONTENT']}")
    print(f"  └ 완전 공백 행        : {total_exclusions['EMPTY_ROW']}")
    print(f"")
    print(f"  서비스별 TC")
    for sname in TARGET_SHEETS:
        if sname in service_stats:
            print(f"    {sname:20s}: {service_stats[sname]['total_tc']} TC")
    print(f"")
    print(f"  Phase 3 품질 이슈: {quality_stats['total_issues']}건  "
          f"(영향 TC: {quality_stats['affected_tc_count']}건)")
    for itype, cnt in quality_stats['issue_type_counts'].items():
        if cnt:
            print(f"    {itype}: {cnt}")
    print(f"{'='*55}")


def _write_attribute_summary(service_stats):
    """
    Phase 2: attribute_summary.csv 생성.
    단일 속성 행 + 복합 속성 조합 행을 함께 기록한다.
    attribute_name 형식:
      단일: 'ui_visibility' (ATTR_KEY 그대로)
      복합: 'compound:UI노출확인+기능동작확인' (compound: 접두사)
    """
    attr_path = os.path.join(OUTPUT_DIR, 'attribute_summary.csv')
    fields = ['service_name', 'attribute_name', 'tc_count', 'ratio']
    total_rows = 0
    with open(attr_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for sname in TARGET_SHEETS:
            if sname not in service_stats:
                continue
            s = service_stats[sname]
            total = s['total_tc']

            # 단일 속성
            a = s['attributes']
            for k in ATTR_KEYS:
                cnt = a.get(k, 0)
                ratio = round(cnt / total, 4) if total else 0.0
                writer.writerow({
                    'service_name'  : sname,
                    'attribute_name': k,
                    'tc_count'      : cnt,
                    'ratio'         : ratio,
                })
                total_rows += 1

            # 복합 속성 조합 (2개 이상 활성 속성을 가진 TC)
            ca = s['compound_attributes']
            for combo, cnt in sorted(ca.items(), key=lambda x: -x[1]):
                ratio = round(cnt / total, 4) if total else 0.0
                writer.writerow({
                    'service_name'  : sname,
                    'attribute_name': f'compound:{combo}',
                    'tc_count'      : cnt,
                    'ratio'         : ratio,
                })
                total_rows += 1

    print(f"[OK] attribute_summary.csv  → {total_rows}행")


def _write_precondition_summary(service_stats, tc_master_rows):
    """
    Phase 4: precondition_summary.csv 생성.

    두 가지 파일 생성:
      A. precondition_summary.csv  — 서비스별 카테고리 분포 (집계)
      B. precondition_detail.csv   — TC별 추론 상세 (근거 포함)
    """
    # ── A. 서비스별 카테고리 분포 ───────────────────────────────────────────
    summary_path = os.path.join(OUTPUT_DIR, 'precondition_summary.csv')
    fields_sum = ['service_name', 'category', 'label', 'tc_count', 'ratio']
    total_rows_a = 0
    with open(summary_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields_sum)
        writer.writeheader()
        for sname in TARGET_SHEETS:
            if sname not in service_stats:
                continue
            s     = service_stats[sname]
            total = s['total_tc']
            pc    = s['precond_categories']
            for rule in PRECOND_RULES:
                cat = rule['category']
                cnt = pc.get(cat, 0)
                if cnt == 0:
                    continue
                writer.writerow({
                    'service_name': sname,
                    'category':     cat,
                    'label':        rule['label'],
                    'tc_count':     cnt,
                    'ratio':        round(cnt / total, 4) if total else 0.0,
                })
                total_rows_a += 1
    print(f"[OK] precondition_summary.csv  → {total_rows_a}행")

    # ── B. TC별 추론 상세 ────────────────────────────────────────────────────
    detail_path = os.path.join(OUTPUT_DIR, 'precondition_detail.csv')
    fields_det = [
        'service_name', 'row_number',
        'category', 'label',
        'evidence', 'source_field', 'keyword', 'reason', 'confidence',
    ]
    total_rows_b = 0
    with open(detail_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields_det)
        writer.writeheader()
        for tc in tc_master_rows:
            infs = infer_preconditions(tc)
            for inf in infs:
                writer.writerow({
                    'service_name': tc['service_name'],
                    'row_number':   tc['row_number'],
                    'category':     inf['category'],
                    'label':        inf['label'],
                    'evidence':     inf['evidence'],
                    'source_field': inf['source_field'],
                    'keyword':      inf['keyword'],
                    'reason':       inf['reason'],
                    'confidence':   inf['confidence'],
                })
                total_rows_b += 1
    print(f"[OK] precondition_detail.csv   → {total_rows_b}행")


# ══════════════════════════════════════════════════════════════════════════════
# Phase 4-B: TC 문맥 분석 및 절차/기대결과 재구성 (규칙 기반)
# ══════════════════════════════════════════════════════════════════════════════

# ── 기대결과 안의 절차(Action) 시그널 ─────────────────────────────────────────
# 패턴 1: "A - B - C" 형태의 메뉴/설정 경로
_PATH_PATTERN = re.compile(
    r'([가-힣\w]+)\s*[-–]\s*([가-힣\w]+)\s*[-–]\s*([가-힣\w]+)',
)
# 패턴 2: 기대결과 문장 끝이 행위 동사로 끝남 (이동, 탭, 클릭, 선택, 진입)
_ACT_END_PATTERN = re.compile(
    r'(탭|클릭|선택|이동|진입|확인\s*후|터치)\s*[.,。]?\s*$',
    re.MULTILINE,
)
# ── 기대결과(State) 시그널 ──────────────────────────────────────────────────
_STATE_PATTERN = re.compile(
    r'(이어야\s*한다|되어야\s*한다|로\s*설정\s*되|이\s*표시|이\s*노출|'
    r'확인된다|표시된다|변경된다|활성화된다|비활성화된다|이동된다|전환된다|'
    r'설정\s*되어\s*있|상태로\s*표시|상태여야|로\s*설정)',
    re.IGNORECASE,
)
# ── Step 안의 확인 행위 시그널 (절차에 검증이 섞인 경우) ───────────────────
_VERIFY_IN_STEP = re.compile(
    r'확인한다\s*$|확인\s*후\s*$|상태\s*확인$',
    re.MULTILINE,
)

# 경로 패턴을 "A > B > C" 형태로 정규화
def _normalize_path(text: str) -> str:
    return _PATH_PATTERN.sub(
        lambda m: f"{m.group(1)} > {m.group(2)} > {m.group(3)}", text
    )

# 기대결과의 각 줄을 절차/기대결과로 분류
def _classify_lines(text: str):
    """
    기대결과 텍스트의 각 줄을 'action' 또는 'expected'로 분류한다.
    반환: list of (line: str, role: 'action'|'expected')
    """
    results = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        # 번호 제거 후 내용만 추출
        content = re.sub(r'^\s*\d+\s*[.)]\s*', '', stripped).strip()
        if not content:
            continue

        has_path   = bool(_PATH_PATTERN.search(content))
        has_act    = bool(_ACT_END_PATTERN.search(content))
        has_state  = bool(_STATE_PATTERN.search(content))

        # 경로 패턴 → 행위로 판단
        if has_path and not has_state:
            results.append((content, 'action'))
        # 행위 동사로 끝나고 상태 표현 없음 → 행위
        elif has_act and not has_state:
            results.append((content, 'action'))
        # 상태 표현 있음 → 기대결과
        elif has_state:
            results.append((content, 'expected'))
        # 판단 불가 → 기대결과로 유지
        else:
            results.append((content, 'expected'))

    return results


def reconstruct_tc(tc_dict: dict) -> dict:
    """
    Phase 4-B: 단일 TC의 절차와 기대결과를 분석·재구성한다.

    반환값:
      {
        'recon_type'      : str,   # 'no_change'|'separated'|'clarified'
        'recon_steps'     : list,  # 재구성된 실행 절차 (번호 없는 텍스트 목록)
        'recon_expected'  : list,  # 재구성된 기대결과 텍스트 목록
        'moved_lines'     : list,  # 기대결과 → 절차로 이동한 줄 목록
        'reason'          : str,   # 재구성 사유 설명
        'evidence'        : dict,  # {'action_signals': [...], 'state_signals': [...]}
      }
    """
    step     = tc_dict.get('test_step', '') or ''
    expected = tc_dict.get('expected_result', '') or ''

    if not expected.strip():
        return {'recon_type': 'no_change', 'recon_steps': [], 'recon_expected': [],
                'moved_lines': [], 'reason': '기대결과가 비어있어 재구성 대상이 아닙니다.',
                'evidence': {}}

    classified = _classify_lines(expected)
    action_lines   = [l for l, r in classified if r == 'action']
    expected_lines = [l for l, r in classified if r == 'expected']

    # 이동할 행이 없으면 변경 없음
    if not action_lines:
        return {'recon_type': 'no_change', 'recon_steps': [], 'recon_expected': [],
                'moved_lines': [], 'reason': '기대결과 내 절차 혼재가 감지되지 않았습니다.',
                'evidence': {}}

    # 기존 step에서 번호 달린 줄 추출
    orig_steps = []
    for line in step.splitlines():
        s = re.sub(r'^\s*\d+\s*[.)]\s*', '', line.strip()).strip()
        if s:
            orig_steps.append(s)

    # 재구성: 기존 절차 + 이동된 절차 행
    recon_steps = orig_steps + [_normalize_path(l) for l in action_lines]

    # 기대결과가 아예 없어지는 경우 → 기대결과=없음으로 표시
    recon_expected = expected_lines if expected_lines else []

    # 재구성 유형 판단
    if expected_lines:
        recon_type = 'separated'   # 절차/기대결과 분리
        reason_parts = []
        for al in action_lines:
            if _PATH_PATTERN.search(al):
                reason_parts.append(
                    f'"{al[:50]}"은 메뉴 경로 탐색 절차이므로 실행 절차로 이동하였습니다.'
                )
            else:
                reason_parts.append(
                    f'"{al[:50]}"은 행위 동사로 끝나는 실행 절차이므로 실행 절차로 이동하였습니다.'
                )
        for el in expected_lines:
            reason_parts.append(
                f'"{el[:50]}"은 최종 상태를 기술하는 기대결과로 유지하였습니다.'
            )
        reason = ' '.join(reason_parts)
    else:
        recon_type = 'clarified'   # 기대결과 전체가 절차였음
        reason = (
            f'기대결과 전체({len(action_lines)}줄)가 실행 절차로 판단되어 '
            f'실행 절차로 이동하였습니다. '
            f'명시적인 기대결과가 없어 별도 기재가 필요합니다.'
        )

    # 근거 시그널
    action_signals = []
    for al in action_lines:
        if _PATH_PATTERN.search(al):
            action_signals.append(f'경로 패턴: {al[:60]}')
        if _ACT_END_PATTERN.search(al):
            action_signals.append(f'행위 동사 종결: {al[:60]}')
    state_signals = [f'상태 표현: {el[:60]}' for el in expected_lines]

    return {
        'recon_type'    : recon_type,
        'recon_steps'   : recon_steps,
        'recon_expected': recon_expected,
        'moved_lines'   : action_lines,
        'reason'        : reason,
        'evidence'      : {
            'action_signals': action_signals,
            'state_signals' : state_signals,
        },
    }


def _write_reconstruction_summary(tc_master_rows: list):
    """
    Phase 4-B: tc_reconstruction.csv 생성.

    재구성이 발생한 TC만 기록한다.
    """
    path = os.path.join(OUTPUT_DIR, 'tc_reconstruction.csv')
    fields = [
        'service_name', 'row_number',
        'recon_type',
        'orig_step', 'orig_expected',
        'recon_steps', 'recon_expected',
        'moved_lines', 'reason',
        'action_signals', 'state_signals',
    ]
    count = 0
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for tc in tc_master_rows:
            result = reconstruct_tc(tc)
            if result['recon_type'] == 'no_change':
                continue
            writer.writerow({
                'service_name'   : tc['service_name'],
                'row_number'     : tc['row_number'],
                'recon_type'     : result['recon_type'],
                'orig_step'      : tc['test_step'][:300],
                'orig_expected'  : tc['expected_result'][:300],
                'recon_steps'    : ' / '.join(result['recon_steps'])[:400],
                'recon_expected' : ' / '.join(result['recon_expected'])[:400],
                'moved_lines'    : ' | '.join(result['moved_lines'])[:300],
                'reason'         : result['reason'][:400],
                'action_signals' : ' | '.join(result['evidence'].get('action_signals', []))[:200],
                'state_signals'  : ' | '.join(result['evidence'].get('state_signals', []))[:200],
            })
            count += 1
    print(f"[OK] tc_reconstruction.csv  → {count}행")
    return count


# ══════════════════════════════════════════════════════════════════════════════
# Phase 4-C: Context Linking (문맥 연계 분석, 규칙 기반)
# ══════════════════════════════════════════════════════════════════════════════

# 생략 패턴 — Step에 이 표현이 있으면 Expected에서 대상을 찾는다
_OMIT_PATTERNS = {
    '다음_화면':  re.compile(r'다음\s*화면', re.I),
    '랜딩_확인':  re.compile(r'랜딩\s*확인', re.I),
    '화면_확인':  re.compile(r'화면\s*확인', re.I),
    '팝업_확인':  re.compile(r'팝업\s*확인', re.I),
    '결과_확인':  re.compile(r'결과\s*확인', re.I),
    '노출_확인':  re.compile(r'노출\s*확인', re.I),
    '동작_확인':  re.compile(r'동작\s*확인', re.I),
    '목록_확인':  re.compile(r'목록\s*확인', re.I),
    'UI_확인':    re.compile(r'UI\s*확인', re.I),
    '상태_확인':  re.compile(r'상태\s*확인', re.I),
    '정상_여부':  re.compile(r'정상\s*여부', re.I),
}

# Expected에서 핵심 명사구(화면명·상태명)를 추출하는 패턴 (우선순위 순)
_SUBJECT_PATS = [
    re.compile(r'([가-힣\w\s()\-/]+?(?:화면|페이지|팝업|뷰|창|모달|시트|목록|리스트))\s*(?:이|가|으로|로)?\s*(?:이동|랜딩|노출|표시|확인)'),
    re.compile(r'([가-힣\w\s()\-/]{2,25})\s*(?:으로|로)\s*이동'),
    re.compile(r'([가-힣\w\s()\-/]{2,25}?)\s*(?:이|가)\s*노출'),
    re.compile(r'([가-힣\w\s()\-/]{2,25}?)\s*(?:이|가)\s*표시'),
    re.compile(r'^([a-zA-Z][a-zA-Z\s/\-]{2,30}(?:menu|view|screen|list))\s', re.I),
]

def _extract_subject_from_expected(exp_text: str) -> str | None:
    """Expected 첫 줄에서 핵심 명사구를 추출한다."""
    first = exp_text.strip().splitlines()[0].strip() if exp_text.strip() else ''
    first = re.sub(r'^\d+\s*[.)]\s*', '', first).strip()
    if not first:
        return None
    # 짧은 문장은 명사구 자체로 간주
    if len(first) <= 20 and not re.search(r'[이가]\s*(노출|표시|이동)', first):
        return first
    for pat in _SUBJECT_PATS:
        m = pat.search(first)
        if m:
            candidate = m.group(1).strip()
            # 너무 짧거나 의미없는 조각은 제외
            if len(candidate) >= 2:
                return candidate[:35]
    return None


def _make_context_linked_step(step: str, omit_key: str, subject: str) -> str:
    """
    생략된 대상이 채워진 재구성 Step을 반환한다.
    omit_key: 생략 패턴 키 (예: '다음_화면')
    subject : Expected에서 추출된 대상 (예: 'MO 인증화면')
    """
    omit_display = omit_key.replace('_', ' ')
    # 생략 패턴을 실제 대상으로 치환
    pat = _OMIT_PATTERNS[omit_key]
    rewritten = pat.sub(subject, step, count=1)
    if rewritten == step:
        # 치환 안 된 경우 보완 문장 추가
        rewritten = step.rstrip() + f'\n  ※ {omit_display} = {subject}'
    return rewritten


def link_context(tc_dict: dict) -> dict:
    """
    Phase 4-C: 단일 TC에서 Step-Expected 간 문맥 연계를 분석한다.

    반환값:
      {
        'linked'         : bool,   # 연계가 발생했는지
        'omit_key'       : str,    # 탐지된 생략 패턴 키
        'omit_expr'      : str,    # 생략 표현 원문
        'subject'        : str,    # Expected에서 추출한 대상
        'source_col'     : str,    # 대상을 찾은 컬럼 ('expected_result' 등)
        'recon_step'     : str,    # 재구성된 Step
        'recon_expected' : str,    # 재구성된 Expected (보완)
        'reason'         : str,    # 판단 근거 설명
        'confidence'     : str,    # 'high'|'medium'|'low'
      }
    """
    step     = tc_dict.get('test_step', '') or ''
    expected = tc_dict.get('expected_result', '') or ''
    precond  = tc_dict.get('precondition', '') or ''

    if not step.strip() or not expected.strip():
        return {'linked': False}

    # Step에서 생략 패턴 탐지
    found_key   = None
    found_match = None
    for key, pat in _OMIT_PATTERNS.items():
        m = pat.search(step)
        if m:
            found_key   = key
            found_match = m.group(0)
            break

    if not found_key:
        return {'linked': False}

    # Expected에서 대상 추출
    subject = _extract_subject_from_expected(expected)
    if not subject:
        return {
            'linked': False,
            'omit_key': found_key, 'omit_expr': found_match,
            'reason': f'Step에 "{found_match}" 생략 표현이 있으나 Expected에서 대상을 추출할 수 없었습니다.',
        }

    # Step 재구성
    recon_step = _make_context_linked_step(step, found_key, subject)

    # Expected 재구성 (명확한 단언문으로)
    first_exp = expected.strip().splitlines()[0]
    first_exp = re.sub(r'^\d+\s*[.)]\s*', '', first_exp).strip()
    # 단언 형태가 없으면 "~이어야 한다" 추가
    if not _STATE_PATTERN.search(first_exp):
        recon_expected = first_exp + ' 이어야 한다.'
    else:
        recon_expected = first_exp

    # Confidence 판단
    omit_display = found_key.replace('_', ' ')
    if found_key in ('다음_화면', '랜딩_확인'):
        confidence = 'high'
    elif len(subject) >= 5:
        confidence = 'high'
    else:
        confidence = 'medium'

    reason = (
        f'Step에 "{found_match}"라는 생략 표현이 있습니다. '
        f'Expected Result의 첫 번째 문장에서 '
        f'"{subject}"이(가) 확인 대상임을 파악하였습니다. '
        f'따라서 "{omit_display}"은 "{subject}"을 의미하는 것으로 판단하였습니다.'
    )

    return {
        'linked'         : True,
        'omit_key'       : found_key,
        'omit_expr'      : found_match,
        'subject'        : subject,
        'source_col'     : 'expected_result',
        'recon_step'     : recon_step,
        'recon_expected' : recon_expected,
        'reason'         : reason,
        'confidence'     : confidence,
    }


def _write_context_linking(tc_master_rows: list):
    """Phase 4-C: tc_context_linking.csv 생성."""
    path = os.path.join(OUTPUT_DIR, 'tc_context_linking.csv')
    fields = [
        'service_name', 'row_number',
        'omit_key', 'omit_expr', 'subject', 'source_col',
        'orig_step', 'orig_expected',
        'recon_step', 'recon_expected',
        'reason', 'confidence',
    ]
    count = 0
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for tc in tc_master_rows:
            result = link_context(tc)
            if not result.get('linked'):
                continue
            writer.writerow({
                'service_name'   : tc['service_name'],
                'row_number'     : tc['row_number'],
                'omit_key'       : result['omit_key'],
                'omit_expr'      : result['omit_expr'],
                'subject'        : result['subject'],
                'source_col'     : result['source_col'],
                'orig_step'      : tc['test_step'][:300],
                'orig_expected'  : tc['expected_result'][:300],
                'recon_step'     : result['recon_step'][:400],
                'recon_expected' : result['recon_expected'][:300],
                'reason'         : result['reason'][:400],
                'confidence'     : result['confidence'],
            })
            count += 1
    print(f"[OK] tc_context_linking.csv  → {count}행")
    return count


# ══════════════════════════════════════════════════════════════════════════════
# 문장 표준화 엔진 (Phase 4-D)
# ══════════════════════════════════════════════════════════════════════════════

# 용어 통일 사전 (원본 → 표준)
_TERM_MAP = [
    # 앱 관련
    (re.compile(r'\b앱을\b'), '애플리케이션을'),
    (re.compile(r'\b앱의\b'), '애플리케이션의'),
    (re.compile(r'\b앱에서\b'), '애플리케이션에서'),
    (re.compile(r'\b앱이\b'), '애플리케이션이'),
    (re.compile(r'\b앱\b'), '애플리케이션'),
    # 동작 표현
    (re.compile(r'화면이 뜬다'), '화면이 표시된다.'),
    (re.compile(r'화면이 뜨면'), '화면이 표시되면'),
    (re.compile(r'뜨는지'), '표시되는지'),
    # 팝업 → 맥락에 따라 유지 (명칭 포함 시 유지)
]

# 사전조건 종결 규칙: (패턴, 치환) 순서 중요
_PRECOND_ENDINGS = [
    (re.compile(r'(?<![.。])\s*로그인\s*(?:완료|되어\s*있음|된\s*상태)?\s*$', re.MULTILINE),
     lambda m: '로그인한 상태이다.'),
    (re.compile(r'(?<![.。])\s*있음\s*$', re.MULTILINE),
     lambda m: '상태이다.'),
    (re.compile(r'(?<![.。])\s*완료\s*$', re.MULTILINE),
     lambda m: '완료되어 있다.'),
    (re.compile(r'(?<![.。])\s*상태\s*$', re.MULTILINE),
     lambda m: '상태이다.'),
    (re.compile(r'(?<![.。다])$', re.MULTILINE),
     lambda m: '이다.'),
]

# Test Step 종결 규칙
_STEP_ENDINGS = [
    (re.compile(r'([^\s])[\s]*탭\s*$', re.MULTILINE),     lambda m: m.group(1) + '을(를) 선택한다.'),
    (re.compile(r'([^\s])[\s]*클릭\s*$', re.MULTILINE),   lambda m: m.group(1) + '을(를) 선택한다.'),
    (re.compile(r'([^\s])[\s]*선택\s*$', re.MULTILINE),   lambda m: m.group(1) + '을(를) 선택한다.'),
    (re.compile(r'([^\s])[\s]*입력\s*$', re.MULTILINE),   lambda m: m.group(1) + '을(를) 입력한다.'),
    (re.compile(r'([^\s])[\s]*진입\s*$', re.MULTILINE),   lambda m: m.group(1) + '에 진입한다.'),
    (re.compile(r'([^\s])[\s]*실행\s*$', re.MULTILINE),   lambda m: m.group(1) + '을(를) 실행한다.'),
    (re.compile(r'([^\s])[\s]*확인\s*$', re.MULTILINE),   lambda m: m.group(1) + '을(를) 확인한다.'),
]

# 기대결과 명사형 → 문장형
_EXP_ENDINGS = [
    (re.compile(r'표시됨\b'), '표시된다.'),
    (re.compile(r'노출됨\b'), '노출된다.'),
    (re.compile(r'이동됨\b'), '이동된다.'),
    (re.compile(r'저장됨\b'), '저장된다.'),
    (re.compile(r'활성화됨\b'), '활성화된다.'),
    (re.compile(r'비활성화됨\b'), '비활성화된다.'),
    (re.compile(r'변경됨\b'), '변경된다.'),
    (re.compile(r'추가됨\b'), '추가된다.'),
    (re.compile(r'삭제됨\b'), '삭제된다.'),
    (re.compile(r'완료됨\b'), '완료된다.'),
    (re.compile(r'생성됨\b'), '생성된다.'),
    (re.compile(r'해제됨\b'), '해제된다.'),
    (re.compile(r'설정됨\b'), '설정된다.'),
    (re.compile(r'전환됨\b'), '전환된다.'),
    (re.compile(r'반영됨\b'), '반영된다.'),
]


def _apply_term_map(text: str) -> tuple:
    """용어 통일 적용. 반환: (변환된 텍스트, 변경 목록)"""
    if not text:
        return text, []
    changes = []
    result = text
    for pat, repl in _TERM_MAP:
        new = pat.sub(repl, result)
        if new != result:
            changes.append(f'용어 통일: {pat.pattern.strip()} → {repl}')
            result = new
    return result, changes


def _std_precond(text: str) -> tuple:
    """사전조건 문장 표준화. 반환: (표준화 텍스트, 변경 목록)"""
    if not text or not text.strip():
        return text, []
    text, changes = _apply_term_map(text)
    lines_out = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            lines_out.append(line)
            continue
        changed = False
        for pat, repl_fn in _PRECOND_ENDINGS:
            new = pat.sub(repl_fn, s)
            if new != s:
                lines_out.append(new)
                changes.append(f'문장 종결 통일 (사전조건): …{s[-8:]} → …{new[-8:]}')
                changed = True
                break
        if not changed:
            lines_out.append(s)
    return '\n'.join(lines_out), changes


def _std_step(text: str) -> tuple:
    """Test Step 문장 표준화. 반환: (표준화 텍스트, 변경 목록)"""
    if not text or not text.strip():
        return text, []
    text, changes = _apply_term_map(text)
    lines_out = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            lines_out.append(line)
            continue
        # 이미 표준 종결이면 건드리지 않음
        if re.search(r'(선택한다|입력한다|실행한다|확인한다|탭한다|진입한다)\.$', s):
            lines_out.append(s)
            continue
        changed = False
        for pat, repl_fn in _STEP_ENDINGS:
            new = pat.sub(repl_fn, s)
            if new != s:
                lines_out.append(new)
                changes.append(f'문장 종결 통일 (Step): …{s[-8:]} → …{new[-8:]}')
                changed = True
                break
        if not changed:
            lines_out.append(s)
    return '\n'.join(lines_out), changes


def _std_expected(text: str) -> tuple:
    """기대결과 문장 표준화. 반환: (표준화 텍스트, 변경 목록)"""
    if not text or not text.strip():
        return text, []
    text, changes = _apply_term_map(text)
    result = text
    for pat, repl in _EXP_ENDINGS:
        new = pat.sub(repl, result)
        if new != result:
            changes.append(f'표현 표준화 (기대결과): {pat.pattern} → {repl}')
            result = new
    return result, changes


def build_norm_summary(pre_changes, step_changes, exp_changes, ctx_changes=None) -> str:
    """Normalization Summary 문자열 생성."""
    all_changes = []
    seen = set()
    for c in (pre_changes + step_changes + exp_changes + (ctx_changes or [])):
        if c not in seen:
            seen.add(c)
            all_changes.append(f'- {c}')
    return '\n'.join(all_changes) if all_changes else ''


# ══════════════════════════════════════════════════════════════════════════════
# Context 추론 엔진 (Phase 4-E)
# ══════════════════════════════════════════════════════════════════════════════

_SCREEN_PATTERN = re.compile(
    r'([가-힣A-Za-z0-9\s\(\)_\-\/]{2,20})'
    r'(?:화면|팝업|페이지|뷰|창|레이어|바텀시트|다이얼로그)',
    re.IGNORECASE,
)

_FLOW_KEYWORDS = {
    '초기': '초기 진입 단계', '최초': '초기 진입 단계', '설치': '초기 진입 단계',
    '오류': '예외 처리 단계', '실패': '예외 처리 단계', '예외': '예외 처리 단계',
    '제한': '예외 처리 단계', '불가': '예외 처리 단계', '차단': '예외 처리 단계',
    '정상': '정상 흐름 단계', '성공': '정상 흐름 단계',
}


def build_context_summary(tc_list: list, idx: int) -> dict:
    """
    서비스 TC 리스트에서 idx 위치 TC의 Context를 추론한다.
    앞뒤 ±5개 TC를 참고한다.
    """
    tc = tc_list[idx]
    window = tc_list[max(0, idx-5): idx+6]

    cat1 = tc.get('category_1', '') or ''
    cat2 = tc.get('category_2', '') or ''
    cat3 = tc.get('category_3', '') or ''
    cat4 = tc.get('category_4', '') or ''
    step = tc.get('test_step', '') or ''
    exp  = tc.get('expected_result', '') or ''

    # ── Feature ────────────────────────────────────────────────────────────
    # cat1이 있으면 사용, 없으면 윈도우에서 가장 빈번한 cat1 사용
    if cat1.strip():
        feature = cat1.strip()
    else:
        from collections import Counter
        cats = [t.get('category_1', '') or '' for t in window if t.get('category_1', '')]
        feature = Counter(cats).most_common(1)[0][0] if cats else ''

    # ── Screen ──────────────────────────────────────────────────────────────
    # step + expected에서 화면명 패턴 추출
    screen = ''
    for text in [step, exp]:
        m = _SCREEN_PATTERN.search(text)
        if m:
            candidate = m.group(0).strip()
            if 3 <= len(candidate) <= 30:
                screen = candidate
                break
    # 없으면 윈도우에서 탐색
    if not screen:
        for t in window:
            for text in [t.get('test_step','') or '', t.get('expected_result','') or '']:
                m = _SCREEN_PATTERN.search(text)
                if m:
                    candidate = m.group(0).strip()
                    if 3 <= len(candidate) <= 30:
                        screen = candidate
                        break
            if screen:
                break

    # ── Scenario ────────────────────────────────────────────────────────────
    # cat2 > cat3 > cat4 순으로 사용
    scenario = cat2.strip() or cat3.strip() or cat4.strip()
    if not scenario and feature:
        scenario = feature

    # ── User Goal ───────────────────────────────────────────────────────────
    # 기대결과 첫 줄을 사용자 관점으로 재작성
    first_exp = (exp or '').split('\n')[0].strip()
    first_exp = re.sub(r'^\d+[.)]\s*', '', first_exp).strip()
    if first_exp:
        # "~표시된다" → "~가 정상 표시되는지 확인"
        user_goal = re.sub(r'(표시된다|노출된다|이동된다|저장된다|활성화된다|비활성화된다|완료된다)\.?$',
                           r'\1 정상 동작하는지 확인한다.', first_exp)
        if user_goal == first_exp:
            user_goal = first_exp + '이(가) 정상 동작하는지 확인한다.'
    else:
        user_goal = ''

    # ── Flow Position ───────────────────────────────────────────────────────
    all_cats = ' '.join(filter(None, [cat1, cat2, cat3, cat4]))
    flow_pos = ''
    for kw, label in _FLOW_KEYWORDS.items():
        if kw in all_cats:
            flow_pos = label
            break
    if not flow_pos:
        # 같은 feature 그룹 내에서 순서 계산
        feature_tcs = [t for t in tc_list if (t.get('category_1','') or '') == cat1] if cat1 else []
        if feature_tcs:
            try:
                pos = feature_tcs.index(tc) + 1
                total = len(feature_tcs)
                flow_pos = f'주요 흐름 {pos}/{total} 단계'
            except ValueError:
                flow_pos = '주요 흐름 단계'
        else:
            flow_pos = '주요 흐름 단계'

    return {
        'ctx_feature'     : feature,
        'ctx_screen'      : screen,
        'ctx_scenario'    : scenario,
        'ctx_user_goal'   : user_goal,
        'ctx_flow_position': flow_pos,
    }


# ── Semantic Validation 헬퍼 ─────────────────────────────────────────────────

_SEM_STOP = {
    '이다','이어야','되어야','한다','있다','없다','된다','이동','이동된','확인',
    '수행','진입','탭','클릭','선택','표시','노출','설정','상태','화면',
    '이상','이하','이후','이전','기준','경우','있는','없는','하는','하여',
    '되는','되어','으로','에서','에게','부터','까지','또는','그리고','및',
}

def _sem_keywords(text: str) -> set:
    import re as _re
    words = _re.findall(r'[가-힣a-zA-Z]{2,}', text or '')
    return {w for w in words if w not in _SEM_STOP}

def _sem_match(orig_step, orig_exp, ai_step, ai_exp) -> int:
    ok = _sem_keywords(orig_step + ' ' + orig_exp)
    ak = _sem_keywords(ai_step  + ' ' + ai_exp)
    if not ok:
        return 100
    return round(len(ok & ak) / len(ok) * 100)

def _build_intent(cat1, cat2, expected) -> str:
    parts = []
    if cat1: parts.append(cat1)
    if cat2 and cat2 != cat1: parts.append(cat2)
    first = (expected or '').split('\n')[0].strip()
    first = re.sub(r'^\d+[.)]\s*', '', first).strip()
    if first and len(first) < 60:
        parts.append(first)
    return ' > '.join(parts)[:100]

def _final_check_point(ai_exp) -> str:
    for line in (ai_exp or '').split('\n'):
        s = re.sub(r'^\d+[.)]\s*', '', line.strip())
        if re.search(r'이어야|되어야|표시|노출|이동|설정', s) and len(s) > 4:
            return s[:80]
    first = (ai_exp or '').split('\n')[0].strip()
    return re.sub(r'^\d+[.)]\s*', '', first)[:80]


def _write_ai_standard_tc(tc_master_rows: list):
    """
    공통 데이터 레이어: ai_standard_tc.csv 생성.
    Dashboard와 Excel Export가 동일 데이터를 사용한다.

    컬럼 그룹:
      A. 원본 식별자: service_name, row_number
      B. Original TC: orig_cat1~orig_expected, orig_priority
      C. AI Standard TC: ai_cat1~ai_expected, ai_priority
      D. AI Analysis:
           precond_norm_type, precond_evidence
           step_norm_type, step_reason
           expected_norm_type, expected_reason
           quality_issues
      E. Intent: original_intent, ai_intent, final_check_point
      F. Semantic Validation: meaning_match_pct, meaning_status, sem_reason
    """
    # ── 보조 데이터 로드 ────────────────────────────────────────────────────
    from collections import defaultdict

    recon_map  = {}
    link_map   = {}
    precond_map = defaultdict(list)
    qi_map     = defaultdict(list)

    recon_path = os.path.join(OUTPUT_DIR, 'tc_reconstruction.csv')
    if os.path.exists(recon_path):
        with open(recon_path, encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                recon_map[r['service_name']+'|'+r['row_number']] = r

    link_path = os.path.join(OUTPUT_DIR, 'tc_context_linking.csv')
    if os.path.exists(link_path):
        with open(link_path, encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                link_map[r['service_name']+'|'+r['row_number']] = r

    pd_path = os.path.join(OUTPUT_DIR, 'precondition_detail.csv')
    if os.path.exists(pd_path):
        with open(pd_path, encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                precond_map[r['service_name']+'|'+r['row_number']].append(r)

    qi_path = os.path.join(OUTPUT_DIR, 'quality_issues.csv')
    if os.path.exists(qi_path):
        with open(qi_path, encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                qi_map[r['service_name']+'|'+r['row_number']].append(r)

    # ── 필드 정의 ────────────────────────────────────────────────────────────
    fields = [
        # 식별자
        'service_name', 'row_number',
        # Original TC
        'orig_cat1', 'orig_cat2', 'orig_cat3', 'orig_cat4',
        'orig_precond', 'orig_step', 'orig_expected', 'orig_priority',
        # AI Standard TC (TC 값만, 설명 없음)
        'ai_cat1', 'ai_cat2', 'ai_cat3', 'ai_cat4',
        'ai_precond', 'ai_step', 'ai_expected', 'ai_priority',
        # AI Analysis
        'precond_norm_type', 'precond_evidence',
        'step_norm_type', 'step_reason',
        'expected_norm_type', 'expected_reason',
        'quality_issues',
        # Intent
        'original_intent', 'ai_intent', 'final_check_point',
        # Semantic Validation
        'meaning_match_pct', 'meaning_status', 'sem_reason',
        # Context Summary (신규)
        'ctx_feature', 'ctx_screen', 'ctx_scenario',
        'ctx_user_goal', 'ctx_flow_position',
        # Normalization Summary (신규)
        'norm_summary',
    ]

    # ── 서비스별 Context 사전 계산 ──────────────────────────────────────────
    from collections import defaultdict as _dd
    svc_tc_map = _dd(list)
    for tc in tc_master_rows:
        svc_tc_map[tc['service_name']].append(tc)

    ctx_cache = {}
    for svc, tc_list in svc_tc_map.items():
        for idx, tc in enumerate(tc_list):
            key = tc['service_name'] + '|' + str(tc['row_number'])
            ctx_cache[key] = build_context_summary(tc_list, idx)

    path = os.path.join(OUTPUT_DIR, 'ai_standard_tc.csv')
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()

        for tc in tc_master_rows:
            svc = tc['service_name']
            num = tc['row_number']
            key = svc + '|' + str(num)

            recon = recon_map.get(key, {})
            link  = link_map.get(key, {})
            pds   = precond_map.get(key, [])
            qis   = qi_map.get(key, [])

            # ── Original TC ────────────────────────────────────────────────
            orig_pre = tc.get('precondition', '') or ''
            orig_stp = tc.get('test_step', '') or ''
            orig_exp = tc.get('expected_result', '') or ''
            orig_pri = tc.get('priority', '') or ''

            # ── AI Standard TC — TC 값만 ────────────────────────────────────
            # 사전조건: 원본 있으면 유지, 없고 추론 있으면 추론 레이블 목록
            if orig_pre.strip():
                ai_pre = orig_pre
            elif pds:
                seen = {}
                labels = [p['label'] for p in pds if not (p['label'] in seen or seen.update({p['label']:1}))]
                ai_pre = '\n'.join(labels[:5])
            else:
                ai_pre = ''

            # Test Step
            recon_type = recon.get('recon_type', '') if recon else ''
            if recon_type in ('separated', 'clarified') and recon.get('recon_steps'):
                ai_stp = recon['recon_steps'].replace(' / ', '\n')
            elif link and link.get('recon_step'):
                ai_stp = link['recon_step']
            else:
                ai_stp = orig_stp

            # 기대결과
            if recon_type == 'separated' and recon.get('recon_expected'):
                ai_exp = recon['recon_expected'].replace(' / ', '\n')
            elif recon_type == 'clarified':
                ai_exp = '(기대결과 재작성 필요)'
            elif link and link.get('recon_expected'):
                ai_exp = link['recon_expected']
            else:
                ai_exp = orig_exp

            # ── 문장 표준화 적용 ────────────────────────────────────────────
            ai_pre_std, pre_changes = _std_precond(ai_pre)
            ai_stp_std, stp_changes = _std_step(ai_stp)
            ai_exp_std, exp_changes = _std_expected(ai_exp)

            # clarified는 표준화 대상 아님 (재작성 필요 메시지 그대로)
            if recon_type != 'clarified':
                ai_pre = ai_pre_std
                ai_stp = ai_stp_std
                ai_exp = ai_exp_std
            else:
                pre_changes, stp_changes, exp_changes = [], [], []

            # ── AI Analysis ─────────────────────────────────────────────────
            # 사전조건 분석
            precond_norm = ''
            precond_ev   = ''
            if pds:
                if not orig_pre.strip():
                    precond_norm = '사전조건 추론'
                    ev_parts = []
                    seen = {}
                    for p in pds[:3]:
                        if p['label'] not in seen:
                            seen[p['label']] = 1
                            ev_parts.append(f"• {p['label']}: {p['reason'][:100]}")
                    precond_ev = '\n'.join(ev_parts)
            if pre_changes:
                precond_norm = ', '.join(filter(None, [precond_norm, '문장 표준화']))

            # Test Step 분석
            step_norm_parts = []
            step_reason = ''
            if recon_type == 'separated':
                step_norm_parts.append('수행 절차 재구성')
                step_reason = recon.get('reason', '')[:200]
            elif recon_type == 'clarified':
                step_norm_parts.append('기대결과 명확화')
                step_reason = recon.get('reason', '')[:200]
            if link and link.get('omit_expr'):
                step_norm_parts.append('문맥 보완')
                if not step_reason:
                    step_reason = link.get('reason', '')[:200]
            if stp_changes:
                step_norm_parts.append('문장 표준화')
            step_norm = ', '.join(step_norm_parts)

            # 기대결과 분석
            exp_norm_parts = []
            exp_reason = ''
            if recon_type == 'separated':
                exp_norm_parts.append('수행 절차 재구성')
            elif recon_type == 'clarified':
                exp_norm_parts.append('기대결과 재작성 필요')
            if link and link.get('recon_expected') and link['recon_expected'] != orig_exp:
                exp_norm_parts.append('표현 표준화')
                if not exp_reason:
                    exp_reason = f'생략된 "{link.get("omit_expr","")}"를 "{link.get("subject","")}"로 보완'
            if exp_changes:
                exp_norm_parts.append('문장 표준화')
            exp_norm = ', '.join(exp_norm_parts)

            # 품질 이슈
            qi_str = '; '.join(
                f'{q["issue_type"]}: {q["issue_reason"][:60]}'
                for q in qis[:3]
            )

            # ── Context Summary ──────────────────────────────────────────────
            ctx = ctx_cache.get(key, {})
            ctx_changes = []
            if ctx.get('ctx_screen'):
                ctx_changes.append(f'Context 기반 화면명: {ctx["ctx_screen"]}')

            # ── Normalization Summary ────────────────────────────────────────
            norm_sum = build_norm_summary(pre_changes, stp_changes, exp_changes, ctx_changes)

            # ── Intent ──────────────────────────────────────────────────────
            orig_intent = _build_intent(tc.get('category_1',''), tc.get('category_2',''), orig_exp)
            ai_intent   = _build_intent(tc.get('category_1',''), tc.get('category_2',''), ai_exp)

            # Final Check Point 강화: 사용자 관점 검증 목적
            fcp_raw = _final_check_point(ai_exp)
            screen  = ctx.get('ctx_screen', '')
            feature = ctx.get('ctx_feature', '')
            if fcp_raw:
                # 화면/기능 context가 있고, fcp_raw가 이미 그 내용을 포함하지 않는 경우만 prefix 추가
                subject_hint = screen or feature
                if subject_hint and subject_hint not in fcp_raw:
                    fcp = f'{subject_hint}에서 {fcp_raw}'
                    # 종결 형태가 없으면 추가
                    if not fcp.rstrip().endswith(('다.', '다', '요.')):
                        fcp += '이(가) 정상 동작하는지 확인한다.'
                else:
                    # 이미 종결 형태면 유지, 아니면 보완
                    if not fcp_raw.rstrip().endswith(('다.', '다', '요.')):
                        fcp = fcp_raw + '이(가) 정상 동작하는지 확인한다.'
                    else:
                        fcp = fcp_raw
            else:
                fcp = ''

            # ── Semantic Validation ─────────────────────────────────────────
            match_pct = _sem_match(orig_stp, orig_exp, ai_stp, ai_exp)
            preserved = match_pct >= 75
            status    = 'Meaning Preserved' if preserved else 'Meaning Changed'

            step_changed = orig_stp.strip() != ai_stp.strip()
            exp_changed  = orig_exp.strip() != ai_exp.strip()
            pre_changed  = orig_pre.strip() != ai_pre.strip()

            reason_parts = []
            if not step_changed and not exp_changed and not pre_changed:
                reason_parts.append('원본과 동일 — Normalization 없음.')
            else:
                if pre_changed:
                    cats = list({p['label'] for p in pds})[:3]
                    reason_parts.append(f'사전조건 추론: {", ".join(cats)}.')
                if step_changed:
                    if recon_type == 'separated':
                        reason_parts.append('기대결과 내 절차를 Test Step으로 분리.')
                    elif recon_type == 'clarified':
                        reason_parts.append('기대결과가 절차로만 구성 → Test Step으로 이동.')
                    elif link:
                        reason_parts.append(f'"{link.get("omit_expr","")}" → "{link.get("subject","")}" 문맥 보완.')
                if exp_changed and recon_type == 'separated':
                    reason_parts.append('기대결과에서 절차를 분리하여 재구성.')
                if match_pct < 75:
                    reason_parts.append('⚠ 의미 변경 가능성 — 검토 필요.')

            sem_reason = ' / '.join(reason_parts) if reason_parts else '—'

            writer.writerow({
                'service_name': svc, 'row_number': num,
                # Original TC
                'orig_cat1': tc.get('category_1',''), 'orig_cat2': tc.get('category_2',''),
                'orig_cat3': tc.get('category_3',''), 'orig_cat4': tc.get('category_4',''),
                'orig_precond': orig_pre, 'orig_step': orig_stp,
                'orig_expected': orig_exp, 'orig_priority': orig_pri,
                # AI Standard TC
                'ai_cat1': tc.get('category_1',''), 'ai_cat2': tc.get('category_2',''),
                'ai_cat3': tc.get('category_3',''), 'ai_cat4': tc.get('category_4',''),
                'ai_precond': ai_pre, 'ai_step': ai_stp,
                'ai_expected': ai_exp, 'ai_priority': orig_pri,
                # AI Analysis
                'precond_norm_type': precond_norm, 'precond_evidence': precond_ev,
                'step_norm_type': step_norm, 'step_reason': step_reason,
                'expected_norm_type': exp_norm, 'expected_reason': exp_reason,
                'quality_issues': qi_str,
                # Intent
                'original_intent': orig_intent, 'ai_intent': ai_intent,
                'final_check_point': fcp,
                # Semantic Validation
                'meaning_match_pct': match_pct, 'meaning_status': status,
                'sem_reason': sem_reason,
                # Context Summary
                'ctx_feature'      : ctx.get('ctx_feature', ''),
                'ctx_screen'       : ctx.get('ctx_screen', ''),
                'ctx_scenario'     : ctx.get('ctx_scenario', ''),
                'ctx_user_goal'    : ctx.get('ctx_user_goal', ''),
                'ctx_flow_position': ctx.get('ctx_flow_position', ''),
                # Normalization Summary
                'norm_summary': norm_sum,
            })

    print(f"[OK] ai_standard_tc.csv  → {len(tc_master_rows)}행")


def _run_quality_check(tc_master_rows):
    """
    Phase 3: tc_master_rows 전체를 순회해 품질 이슈를 탐지한다.
    반환값:
      quality_issues : list of dict (quality_issues.csv 행)
      quality_stats  : dict (coverage.json quality 블록용)
    """
    quality_issues   = []
    svc_issue_counts = defaultdict(int)
    svc_affected_tcs = defaultdict(set)
    svc_type_counts  = defaultdict(lambda: defaultdict(int))  # [svc][itype]
    type_counts      = defaultdict(int)
    affected_rows    = set()

    for tc in tc_master_rows:
        svc    = tc['service_name']
        rownum = tc['row_number']
        cat_path = ' > '.join(filter(None, [
            tc.get('category_1',''), tc.get('category_2',''),
            tc.get('category_3',''), tc.get('category_4',''),
        ]))

        found = detect_quality_issues(tc)
        for itype, reason in found:
            quality_issues.append({
                'service_name'   : svc,
                'row_number'     : rownum,
                'issue_type'     : itype,
                'issue_reason'   : reason,
                'priority'       : tc['priority'],
                'os'             : tc['os'],
                'category_path'  : cat_path,
                'test_step'      : tc['test_step'][:200],
                'expected_result': tc['expected_result'][:200],
            })
            svc_issue_counts[svc]      += 1
            svc_type_counts[svc][itype] += 1
            type_counts[itype]         += 1
            svc_affected_tcs[svc].add((svc, rownum))
            affected_rows.add((svc, rownum))

    top_itype = max(type_counts, key=type_counts.get) if type_counts else ''
    top_svc   = max(svc_issue_counts, key=svc_issue_counts.get) if svc_issue_counts else ''

    quality_stats = {
        'total_issues'       : len(quality_issues),
        'affected_tc_count'  : len(affected_rows),
        'issue_type_counts'  : {k: type_counts.get(k, 0) for k in ISSUE_TYPES},
        'service_issue_counts': dict(svc_issue_counts),
        'service_affected_tc_counts': {
            svc: len(rows) for svc, rows in svc_affected_tcs.items()
        },
        'svc_type_counts'    : {svc: dict(d) for svc, d in svc_type_counts.items()},
        'top_issue_type'     : top_itype,
        'top_issue_service'  : top_svc,
    }
    return quality_issues, quality_stats


def _write_quality_issues(quality_issues):
    """Phase 3: quality_issues.csv 생성."""
    fields = [
        'service_name', 'row_number', 'issue_type', 'issue_reason',
        'priority', 'os', 'category_path', 'test_step', 'expected_result',
    ]
    path = os.path.join(OUTPUT_DIR, 'quality_issues.csv')
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(quality_issues)
    print(f"[OK] quality_issues.csv  → {len(quality_issues)}행")


def _patch_coverage_quality(cov_path, quality_stats, service_stats):
    """Phase 3: coverage.json을 읽어 quality 블록을 추가/갱신 후 재저장."""
    with open(cov_path, encoding='utf-8') as f:
        coverage = json.load(f)

    # 최상위 quality 블록
    coverage['quality'] = {
        'total_issues'      : quality_stats['total_issues'],
        'affected_tc_count' : quality_stats['affected_tc_count'],
        'issue_type_counts' : quality_stats['issue_type_counts'],
        'service_issue_counts': {
            sname: quality_stats['service_issue_counts'].get(sname, 0)
            for sname in TARGET_SHEETS
        },
        'top_issue_type'    : quality_stats['top_issue_type'],
        'top_issue_service' : quality_stats['top_issue_service'],
    }

    # 각 service 객체에 quality 추가 (issue_type_counts 포함)
    svc_affected   = quality_stats.get('service_affected_tc_counts', {})
    svc_type_dict  = quality_stats.get('svc_type_counts', {})

    for svc_obj in coverage['services']:
        sname = svc_obj['service_name']
        svc_obj['quality'] = {
            'issue_count'      : quality_stats['service_issue_counts'].get(sname, 0),
            'affected_tc_count': svc_affected.get(sname, 0),
            'issue_type_counts': {k: svc_type_dict.get(sname, {}).get(k, 0)
                                  for k in ISSUE_TYPES},
        }

    with open(cov_path, 'w', encoding='utf-8') as f:
        json.dump(coverage, f, ensure_ascii=False, indent=2)
    print(f"[OK] coverage.json  (quality 블록 추가)")


def _write_summary(service_stats, total_tc, quality_stats=None, excl_report=None):
    lines = [
        '# QA Regression TestCase 분석 요약 — Phase 3',
        '',
        f'**분석 대상**: {", ".join(TARGET_SHEETS)}',
        f'**전체 TC 수**: {total_tc}',
        '',
        '---',
        '',
        '## 서비스별 TC 수',
        '',
        '| 서비스 | 전체 TC |',
        '|--------|---------|',
    ]
    for sname in TARGET_SHEETS:
        if sname in service_stats:
            lines.append(f'| {sname} | {service_stats[sname]["total_tc"]} |')

    lines += ['', '## Priority 분포', '']
    p_labels = ['P0', 'P1', 'P2', 'P3', 'P4', 'P5', 'empty', 'other']
    header_row = '| 서비스 | ' + ' | '.join(p_labels) + ' |'
    sep_row    = '|--------|' + '|'.join(['-----'] * len(p_labels)) + '|'
    lines += [header_row, sep_row]
    for sname in TARGET_SHEETS:
        if sname not in service_stats:
            continue
        p = service_stats[sname]['priority']
        vals = ' | '.join(str(p.get(lbl, 0)) for lbl in p_labels)
        lines.append(f'| {sname} | {vals} |')

    lines += ['', '## OS 분포', '']
    os_labels = ['android', 'ios', 'common', 'web', 'pc', 'empty', 'other']
    header_row = '| 서비스 | ' + ' | '.join(os_labels) + ' |'
    sep_row    = '|--------|' + '|'.join(['-----'] * len(os_labels)) + '|'
    lines += [header_row, sep_row]
    for sname in TARGET_SHEETS:
        if sname not in service_stats:
            continue
        o = service_stats[sname]['os']
        vals = ' | '.join(str(o.get(lbl, 0)) for lbl in os_labels)
        lines.append(f'| {sname} | {vals} |')

    attr_labels = {
        'ui_visibility'     : 'UI 노출 확인',
        'data_change'       : '데이터 변경 확인',
        'function_behavior' : '기능 동작 확인',
        'permission_auth'   : '권한/인증 확인',
        'exception_error'   : '예외/에러 확인',
        'network_server'    : '네트워크/서버 연동',
        'notification'      : '알림/푸시 확인',
        'multi_device_os'   : '멀티디바이스/OS 차이',
        'state_persistence' : '설정/상태 유지 확인',
        'content_media'     : '콘텐츠/미디어 확인',
        'has_precondition'  : '사전조건 있음',
        'automation_candidate': '자동화 후보',
        'manual_required'   : '수동 확인 필요',
    }
    lines += ['', '## TC 속성 분포 (Phase 2)', '']
    # 헤더
    attr_header = '| 속성 | ' + ' | '.join(TARGET_SHEETS) + ' |'
    attr_sep    = '|------' + ('|------' * len(TARGET_SHEETS)) + '|'
    lines += [attr_header, attr_sep]
    for k in ATTR_KEYS:
        row_vals = []
        for sname in TARGET_SHEETS:
            if sname not in service_stats:
                row_vals.append('-')
                continue
            a = service_stats[sname]['attributes']
            t = service_stats[sname]['total_tc']
            cnt = a.get(k, 0)
            pct = f'{cnt/t*100:.0f}%' if t else '0%'
            row_vals.append(f'{cnt} ({pct})')
        lines.append(f'| {attr_labels[k]} | ' + ' | '.join(row_vals) + ' |')

    lines += [
        '',
        '## 컬럼 매핑 (3개 시트 공통)',
        '',
        '| 컬럼 인덱스 | 헤더명 | 비고 |',
        '|------------|--------|------|',
        '| 0 | 분류1 | Forward Fill 적용 |',
        '| 1 | 분류2 | Forward Fill 적용 |',
        '| 2 | 분류3 | Forward Fill 적용 |',
        '| 3 | 분류4 (기능) | Forward Fill 적용 |',
        '| 4 | 사전 조건 | |',
        '| 5 | Test Step (체크할 항목) | |',
        '| 6 | 기대결과 | |',
        '| 27 | Priority (P) | 0~5 정수, `-`=비어있음, 기타=집계행 |',
        '| 28 | OS Category (C) | A=Android, I/iOS=iOS, app=공통 |',
        '| 29 | Version (V) | 버전 문자열 |',
        '',
        '## TC 식별 기준',
        '',
        '- Priority(col27) 값이 float로 파싱 가능하고 정수값 0~5 범위인 행 → TC',
        '- `-` / 문자열 / 소수값 / 5 초과 숫자 → 섹션 헤더 또는 집계 행으로 제외',
        '',
        '---',
        '',
        '*생성 파일: tc_master.csv / service_summary.csv / category_summary.csv / coverage.json / '
        'attribute_summary.csv / quality_issues.csv / '
        'precondition_summary.csv / precondition_detail.csv / '
        'tc_reconstruction.csv / summary.md*',
    ]

    # ── TC 판별 통계 섹션 ──────────────────────────────────────────────────────
    if excl_report:
        tr  = excl_report.get('total_read', 0)
        ttc = excl_report.get('total_tc', 0)
        tex = tr - ttc
        lines += [
            '',
            '## TC 판별 통계',
            '',
            f'| 항목 | 행 수 |',
            f'|------|-------|',
            f'| 전체 읽은 행 수 | {tr} |',
            f'| **TC 판정 행 수** | **{ttc}** |',
            f'| 제외 행 수 (합계) | {tex} |',
            f'| ┌ 통계 영역 경계(행 추가시…) 이후 | {excl_report.get("STAT_BOUNDARY", 0)} |',
            f'| ├ 통계표 키워드 행 (OS/Browser, Pass Rate 등) | {excl_report.get("STAT_ROW", 0)} |',
            f'| ├ Priority 없음 / `-` | {excl_report.get("NO_PRIORITY", 0)} |',
            f'| ├ Priority 비허용값 | {excl_report.get("INVALID_PRIORITY", 0)} |',
            f'| ├ 기대결과 공란 (설명·메모 행) | {excl_report.get("NO_EXPECTED", 0)} |',
            f'| ├ Step·분류 모두 없음 | {excl_report.get("NO_CONTENT", 0)} |',
            f'| └ 완전 공백 행 | {excl_report.get("EMPTY_ROW", 0)} |',
            '',
            '> **TC 판별 기준**: Priority(col27) 정수 0~5 AND 기대결과(col6) 비어있지 않음',
            '> 통계 영역 경계("행 추가시…") 이후 행은 모두 분석 제외',
        ]

    # ── Phase 3: 품질 이슈 섹션 ────────────────────────────────────────────────
    if quality_stats:
        issue_label = {
            'ambiguous_test_step'       : 'Test Step 모호',
            'ambiguous_expected_result' : '기대결과 모호',
            'missing_precondition'      : '사전조건 누락',
            'duplicated_step_expected'  : 'Step=Expected 중복',
            'multiple_purpose'          : '복수 목적 혼재',
            'invalid_priority'          : 'Priority 이상값',
            'invalid_os'                : 'OS 이상값',
            'invalid_category'          : '분류 누락/불일치',
        }
        lines += [
            '',
            '## 품질 점검 결과 (Phase 3)',
            '',
            f'- **총 이슈 수**: {quality_stats["total_issues"]}건',
            f'- **영향 TC 수**: {quality_stats["affected_tc_count"]}건 '
            f'(전체 {total_tc}건의 '
            f'{quality_stats["affected_tc_count"]/total_tc*100:.1f}%)',
            f'- **최다 이슈 유형**: {quality_stats["top_issue_type"]}',
            f'- **최다 이슈 서비스**: {quality_stats["top_issue_service"]}',
            '',
            '### 이슈 유형별 건수',
            '',
            '| 이슈 유형 | 건수 |',
            '|-----------|------|',
        ]
        for k in ISSUE_TYPES:
            cnt = quality_stats['issue_type_counts'].get(k, 0)
            lines.append(f'| {issue_label[k]} | {cnt} |')

        lines += [
            '',
            '### 서비스별 품질 이슈 건수',
            '',
            '| 서비스 | 이슈 건수 | 영향 TC 수 |',
            '|--------|-----------|------------|',
        ]
        svc_affected = quality_stats.get('service_affected_tc_counts', {})
        for sname in TARGET_SHEETS:
            ic  = quality_stats['service_issue_counts'].get(sname, 0)
            atc = svc_affected.get(sname, 0)
            lines.append(f'| {sname} | {ic} | {atc} |')

        lines += [
            '',
            '### quality_issues.csv 컬럼 구조',
            '',
            '| 컬럼명 | 설명 |',
            '|--------|------|',
            '| service_name | 서비스명 |',
            '| row_number | Excel 행 번호 |',
            '| issue_type | 이슈 유형 키 |',
            '| issue_reason | 탐지 근거 메시지 |',
            '| priority | 정규화된 Priority |',
            '| os | 정규화된 OS |',
            '| category_path | 분류1 > 2 > 3 > 4 경로 |',
            '| test_step | Test Step 텍스트 (최대 200자) |',
            '| expected_result | 기대결과 텍스트 (최대 200자) |',
        ]

    # ── Phase 4: 사전조건 분석 섹션 ──────────────────────────────────────────────
    # 전체 카테고리 분포 (서비스 합산)
    from collections import Counter as _Counter
    total_precond_cat: dict = _Counter()
    for sname in TARGET_SHEETS:
        if sname not in service_stats:
            continue
        for cat, cnt in service_stats[sname]['precond_categories'].items():
            total_precond_cat[cat] += cnt

    if total_precond_cat:
        cat_label_map = {r['category']: r['label'] for r in PRECOND_RULES}
        lines += [
            '',
            '## 사전조건 속성 분포 (Phase 4)',
            '',
            '| 사전조건 카테고리 | 전체 건수 |',
            '|------------------|-----------|',
        ]
        for cat, cnt in sorted(total_precond_cat.items(), key=lambda x: -x[1]):
            lines.append(f'| {cat_label_map.get(cat, cat)} | {cnt} |')

        lines += [
            '',
            '### 서비스별 상위 사전조건 카테고리',
            '',
        ]
        for sname in TARGET_SHEETS:
            if sname not in service_stats:
                continue
            pc = service_stats[sname]['precond_categories']
            if not pc:
                continue
            top = sorted(pc.items(), key=lambda x: -x[1])[:5]
            lines.append(f'**{sname}**')
            for cat, cnt in top:
                lines.append(f'  - {cat_label_map.get(cat, cat)}: {cnt}건')
            lines.append('')

    summary_path = os.path.join(OUTPUT_DIR, 'summary.md')
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')
    print(f"[OK] summary.md")


if __name__ == '__main__':
    main()
