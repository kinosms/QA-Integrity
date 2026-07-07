#!/usr/bin/env python3
"""
Integrity Dashboard — AI Standard TC Excel Export
output/ai_review.xlsx 생성

공통 데이터 레이어: output/ai_standard_tc.csv

열 구성:
  A~H  : Original TC
  J~R  : AI Standard TC  (TC 값만, 설명 없음)
  T~Z  : AI Analysis     (Normalization Type, Evidence, Reason, Quality Issues)
  AB~AE: Intent          (Original Intent, AI Intent, Final Check Point)
  AG~AI: Semantic Validation
"""

import csv
import os
from collections import defaultdict, OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

# ── 컬러 팔레트 (다크 테마) ────────────────────────────────────────────────
BG_HEADER_ORIG = "1e2235"
BG_HEADER_AI   = "0d1f0d"
BG_HEADER_ANA  = "1a1227"
BG_HEADER_INT  = "111827"
BG_HEADER_SEM  = "0f172a"

BG_ORIG = "12151f"
BG_AI   = "0a120a"
BG_ANA  = "13102a"
BG_INT  = "0c1020"
BG_SEM  = "0a0f1e"

# diff 색상 (배경)
BG_ADDED    = "052e16"   # 초록
BG_MODIFIED = "422006"   # 노랑
BG_DELETED  = "450a0a"   # 빨강
BG_INFERRED = "2e1065"   # 보라 (추론)

# Semantic 색상
BG_SEM_HIGH = "14532d"
BG_SEM_MID  = "713f12"
BG_SEM_LOW  = "7f1d1d"

# 폰트 색상
FC_BRIGHT  = "e2e8f0"
FC_MUTED   = "94a3b8"
FC_ORIG    = "64748b"
FC_GREEN   = "4ade80"
FC_YELLOW  = "fbbf24"
FC_RED     = "f87171"
FC_PURPLE  = "c084fc"
FC_ACCENT  = "818cf8"

def fill(hex_bg):
    return PatternFill("solid", fgColor=hex_bg)

def font(color=FC_BRIGHT, bold=False, size=9, italic=False):
    return Font(color=color, bold=bold, size=size, italic=italic)

ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="2e3350")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── 열 레이아웃 ──────────────────────────────────────────────────────────────
#
# A  분류1 (Orig)    J  분류1 (AI)      T  사전조건 분석    AB  Original Intent
# B  분류2           K  분류2            U  Test Step 분석  AC  AI Intent
# C  분류3           L  분류3            V  기대결과 분석   AD  Final Check Point
# D  화면전개         M  화면전개          W  품질 이슈
# E  사전조건         N  사전조건
# F  Test Step       O  Test Step
# G  기대결과         P  기대결과         AG  Match %
# H  Priority        Q  Priority        AH  Meaning Status
# I  (구분)          R  (구분)           AI  판단 근거
#                    S  (구분)

COL = {
    # Original TC
    'o_cat1':1,'o_cat2':2,'o_cat3':3,'o_cat4':4,
    'o_pre':5,'o_step':6,'o_exp':7,'o_pri':8,
    'sep1':9,
    # AI Standard TC
    'a_cat1':10,'a_cat2':11,'a_cat3':12,'a_cat4':13,
    'a_pre':14,'a_step':15,'a_exp':16,'a_pri':17,
    'sep2':18,'sep3':19,
    # AI Analysis
    'an_pre':20,'an_step':21,'an_exp':22,'an_qi':23,
    'sep4':24,'sep5':25,
    # Intent
    'i_orig':26,'i_ai':27,'i_fcp':28,
    'sep6':29,'sep7':30,
    # Semantic Validation
    's_match':31,'s_status':32,'s_reason':33,
}

WIDTHS = {
    1:16,2:16,3:16,4:16,5:28,6:38,7:38,8:8,
    9:1,
    10:16,11:16,12:16,13:16,14:28,15:38,16:38,17:8,
    18:1,19:1,
    20:35,21:35,22:35,23:30,
    24:1,25:1,
    26:32,27:32,28:32,
    29:1,30:1,
    31:10,32:16,33:35,
}

# ── diff 판단 ─────────────────────────────────────────────────────────────────

def diff_fill(orig, ai, norm_type=''):
    o = (orig or '').strip()
    a = (ai   or '').strip()
    if o == a:            return fill(BG_AI)
    if not o and a:
        if '추론' in norm_type:  return fill(BG_INFERRED)
        return fill(BG_ADDED)
    if o and not a:       return fill(BG_DELETED)
    return fill(BG_MODIFIED)

def diff_font(orig, ai, norm_type=''):
    o = (orig or '').strip()
    a = (ai   or '').strip()
    if o == a:            return font(FC_BRIGHT)
    if not o and a:
        if '추론' in norm_type:  return font(FC_PURPLE, bold=True)
        return font(FC_GREEN, bold=True)
    if o and not a:       return font(FC_RED)
    return font(FC_YELLOW, bold=True)

# ── 셀 쓰기 ──────────────────────────────────────────────────────────────────

def wc(ws, row, col, val, fnt=None, fl=None, align=None):
    c = ws.cell(row=row, column=col, value=val or '')
    c.font      = fnt   or font()
    c.fill      = fl    or fill(BG_ORIG)
    c.alignment = align or ALIGN_WRAP
    c.border    = BORDER
    return c

# ── 시트 구성 ─────────────────────────────────────────────────────────────────

def build_sheet(ws, rows):
    ws.sheet_view.showGridLines = False

    # ── 행 1: 그룹 헤더 ───────────────────────────────────────────────────────
    groups = [
        (1,8,   BG_HEADER_ORIG, "Original TC"),
        (10,17, BG_HEADER_AI,   "AI Standard TC"),
        (20,23, BG_HEADER_ANA,  "AI Analysis"),
        (26,28, BG_HEADER_INT,  "Intent"),
        (31,33, BG_HEADER_SEM,  "Semantic Validation"),
    ]
    for start, end, bg, label in groups:
        fl = fill(bg)
        for c in range(start, end+1):
            ws.cell(row=1, column=c).fill = fl
            ws.cell(row=1, column=c).border = BORDER
        ws.cell(row=1, column=start, value=label)
        ws.cell(row=1, column=start).font = font(FC_BRIGHT, bold=True, size=10)
        ws.cell(row=1, column=start).alignment = ALIGN_CENTER
        if start != end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    # ── 행 2: 컬럼 헤더 ───────────────────────────────────────────────────────
    hdrs = {
        1:'분류1',2:'분류2',3:'분류3',4:'화면전개',
        5:'사전조건',6:'Test Step',7:'기대결과',8:'Priority',
        10:'분류1',11:'분류2',12:'분류3',13:'화면전개',
        14:'사전조건',15:'Test Step',16:'기대결과',17:'Priority',
        20:'사전조건 분석',21:'Test Step 분석',22:'기대결과 분석',23:'품질 이슈',
        26:'Original Intent',27:'AI Intent',28:'Final Check Point',
        31:'Match %',32:'Meaning Status',33:'판단 근거',
    }
    bg_map = {**{c:BG_HEADER_ORIG for c in range(1,9)},
              **{c:BG_HEADER_AI   for c in range(10,18)},
              **{c:BG_HEADER_ANA  for c in range(20,24)},
              **{c:BG_HEADER_INT  for c in range(26,29)},
              **{c:BG_HEADER_SEM  for c in range(31,34)}}
    for col, label in hdrs.items():
        c = ws.cell(row=2, column=col, value=label)
        c.font      = font(FC_BRIGHT, bold=True, size=9)
        c.fill      = fill(bg_map.get(col, BG_HEADER_ORIG))
        c.alignment = ALIGN_CENTER
        c.border    = BORDER

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 22

    # ── 데이터 행 ─────────────────────────────────────────────────────────────
    for data_row, r in enumerate(rows, start=3):
        # Original TC
        wc(ws,data_row,1, r['orig_cat1'],  font(FC_ORIG), fill(BG_ORIG))
        wc(ws,data_row,2, r['orig_cat2'],  font(FC_ORIG), fill(BG_ORIG))
        wc(ws,data_row,3, r['orig_cat3'],  font(FC_ORIG), fill(BG_ORIG))
        wc(ws,data_row,4, r['orig_cat4'],  font(FC_ORIG), fill(BG_ORIG))
        wc(ws,data_row,5, r['orig_precond'],  font(FC_ORIG), fill(BG_ORIG))
        wc(ws,data_row,6, r['orig_step'],     font(FC_ORIG), fill(BG_ORIG))
        wc(ws,data_row,7, r['orig_expected'], font(FC_ORIG), fill(BG_ORIG))
        wc(ws,data_row,8, r['orig_priority'], font(FC_ORIG), fill(BG_ORIG), ALIGN_CENTER)

        # AI Standard TC — TC 값만
        pre_nt = r.get('precond_norm_type','')
        stp_nt = r.get('step_norm_type','')
        exp_nt = r.get('expected_norm_type','')

        wc(ws,data_row,10, r['ai_cat1'], font(FC_BRIGHT), fill(BG_AI))
        wc(ws,data_row,11, r['ai_cat2'],
           diff_font(r['orig_cat2'], r['ai_cat2']),
           diff_fill(r['orig_cat2'], r['ai_cat2']))
        wc(ws,data_row,12, r['ai_cat3'],
           diff_font(r['orig_cat3'], r['ai_cat3']),
           diff_fill(r['orig_cat3'], r['ai_cat3']))
        wc(ws,data_row,13, r['ai_cat4'], font(FC_BRIGHT), fill(BG_AI))
        wc(ws,data_row,14, r['ai_precond'],
           diff_font(r['orig_precond'], r['ai_precond'], pre_nt),
           diff_fill(r['orig_precond'], r['ai_precond'], pre_nt))
        wc(ws,data_row,15, r['ai_step'],
           diff_font(r['orig_step'], r['ai_step'], stp_nt),
           diff_fill(r['orig_step'], r['ai_step'], stp_nt))
        wc(ws,data_row,16, r['ai_expected'],
           diff_font(r['orig_expected'], r['ai_expected'], exp_nt),
           diff_fill(r['orig_expected'], r['ai_expected'], exp_nt))
        wc(ws,data_row,17, r['ai_priority'], font(FC_BRIGHT), fill(BG_AI), ALIGN_CENTER)

        # AI Analysis — 분석 내용만
        def analysis_cell(col, norm_type, evidence, reason):
            parts = []
            if norm_type:
                parts.append(f'[Normalization Type]\n{norm_type}')
            if evidence:
                parts.append(f'[근거]\n{evidence}')
            if reason:
                parts.append(f'[수정 이유]\n{reason}')
            txt = '\n\n'.join(parts)
            wc(ws, data_row, col, txt, font(FC_BRIGHT, size=8), fill(BG_ANA))

        analysis_cell(20,
            r.get('precond_norm_type',''),
            r.get('precond_evidence',''),
            '')
        analysis_cell(21,
            r.get('step_norm_type',''),
            '',
            r.get('step_reason',''))
        analysis_cell(22,
            r.get('expected_norm_type',''),
            '',
            r.get('expected_reason',''))
        wc(ws,data_row,23, r.get('quality_issues',''),
           font(FC_RED if r.get('quality_issues') else FC_MUTED, size=8),
           fill(BG_ANA))

        # Intent
        wc(ws,data_row,26, r.get('original_intent',''), font(FC_MUTED, size=9), fill(BG_INT))
        wc(ws,data_row,27, r.get('ai_intent',''),       font(FC_BRIGHT,size=9), fill(BG_INT))
        wc(ws,data_row,28, r.get('final_check_point',''),font(FC_BRIGHT,bold=True,size=9), fill(BG_INT))

        # Semantic Validation
        match_pct = int(r.get('meaning_match_pct', 100) or 100)
        match_bg  = BG_SEM_HIGH if match_pct >= 90 else BG_SEM_MID if match_pct >= 75 else BG_SEM_LOW
        wc(ws,data_row,31, f'{match_pct}%',
           font(FC_BRIGHT, bold=True, size=11),
           fill(match_bg), ALIGN_CENTER)

        status = r.get('meaning_status','')
        s_color = FC_GREEN if 'Preserved' in status else FC_RED
        wc(ws,data_row,32,
           ('✔ ' if 'Preserved' in status else '✘ ') + status,
           font(s_color, bold=True, size=9),
           fill(BG_SEM), ALIGN_CENTER)

        wc(ws,data_row,33, r.get('sem_reason',''), font(FC_MUTED, size=8), fill(BG_SEM))

    # ── 열 너비 & 구분 열 ──────────────────────────────────────────────────────
    for col, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for sep in [9,18,19,24,25,29,30]:
        ws.column_dimensions[get_column_letter(sep)].width = 0.8
        for row_idx in range(1, len(rows)+3):
            c = ws.cell(row=row_idx, column=sep)
            c.fill = fill("0a0c15")

    ws.freeze_panes = 'A3'


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("ai_standard_tc.csv 로딩 중...")
    rows_by_svc = OrderedDict()
    with open(os.path.join(OUTPUT_DIR, 'ai_standard_tc.csv'), encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            svc = r['service_name']
            if svc not in rows_by_svc:
                rows_by_svc[svc] = []
            rows_by_svc[svc].append(r)

    total = sum(len(v) for v in rows_by_svc.values())
    print(f"서비스: {len(rows_by_svc)}개, 전체 TC: {total}건")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for svc, rows in rows_by_svc.items():
        ws = wb.create_sheet(title=svc[:31])
        ws.sheet_properties.tabColor = "1e2235"
        build_sheet(ws, rows)
        print(f"  [{svc}] {len(rows)}건")

    out = os.path.join(OUTPUT_DIR, 'ai_review.xlsx')
    wb.save(out)
    print(f"\n[OK] {out} ({os.path.getsize(out)/1024/1024:.1f} MB)")


if __name__ == '__main__':
    main()
