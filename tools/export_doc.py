#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_doc.py — 변환 결과를 서식 있는 Excel 정의서로 출력

탭 구성 (요청 순서):
  1. Screen Dictionary
  2. Target Dictionary
  3. State Dictionary
  4. Transition Dictionary
  5. 5.일반채팅 TC 리스트   (신규 구조화 TC + Step)
  6. 변환 요약

입력:  output/chat_conversion/conversion.json
출력:  output/chat_conversion/일반채팅_구조화TC_정의서.xlsx

실행:  python tools/export_doc.py
"""

import os
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE_DIR, 'output', 'chat_conversion', 'conversion.json')
OUT = os.path.join(BASE_DIR, 'output', 'chat_conversion', '일반채팅_구조화TC_정의서.xlsx')

# ── 스타일 ──────────────────────────────────────────────────────────────────
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
HEAD_FONT = Font(color='FFFFFF', bold=True, size=10)
TITLE_FONT = Font(color='1F3864', bold=True, size=14)
SUB_FONT = Font(color='595959', size=9, italic=True)
GROUP_FILL_A = PatternFill('solid', fgColor='FFFFFF')
GROUP_FILL_B = PatternFill('solid', fgColor='EEF3FB')
NA_FONT = Font(color='A6A6A6', size=9)
CELL_FONT = Font(size=9)
MONO_FONT = Font(color='2F5496', size=9, bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical='top', wrap_text=True)
CENTER = Alignment(vertical='center', horizontal='center')


def _sheet(wb, title, subtitle, headers, widths):
    ws = wb.create_sheet(title[:31])
    # 제목
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(2, 1, subtitle).font = SUB_FONT
    # 헤더 (4행)
    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        cell.border = BORDER
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[hr].height = 24
    ws.freeze_panes = ws.cell(hr + 1, 1)
    return ws, hr + 1


def _row(ws, r, values, mono_cols=(), na_check=True):
    for c, v in enumerate(values, 1):
        cell = ws.cell(r, c, v)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if c in mono_cols and isinstance(v, str) and v and v != '해당없음':
            cell.font = MONO_FONT
        elif na_check and v == '해당없음':
            cell.font = NA_FONT
        else:
            cell.font = CELL_FONT


def build():
    with open(SRC, encoding='utf-8') as f:
        D = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. Screen Dictionary ────────────────────────────────────────────────
    ws, r = _sheet(wb, 'Screen Dictionary', '화면 위치·식별 기준 정의서 (5.일반채팅)',
                   ['Screen ID', '화면명', '식별 기준', '출처 원본 행번호', '사용 TC 수'],
                   [26, 22, 46, 40, 11])
    for s in D['screens']:
        _row(ws, r, [s['screen_id'], s['name'], s['criteria'],
                     ' '.join(map(str, s['source_rows'])), s['used_tc_count']], mono_cols=(1,))
        r += 1

    # ── 2. Target Dictionary ────────────────────────────────────────────────
    ws, r = _sheet(wb, 'Target Dictionary', '조작·검증 대상 식별 기준 정의서 (5.일반채팅)',
                   ['Target ID', '대상명', 'Screen ID', '식별 기준', '출처 원본 행번호', '사용 TC 수'],
                   [30, 20, 26, 40, 34, 11])
    for t in D['targets']:
        _row(ws, r, [t['target_id'], t['name'], t['screen_id'], t['criteria'],
                     ' '.join(map(str, t['source_rows'])), t['used_tc_count']], mono_cols=(1, 3))
        r += 1

    # ── 3. State Dictionary ─────────────────────────────────────────────────
    ws, r = _sheet(wb, 'State Dictionary', '재사용 가능한 공통 상태 정의서 (5.일반채팅)',
                   ['State ID', 'State Type', '정의', '판정 기준', '출처 원본 행번호', '사용 TC 수'],
                   [32, 16, 40, 42, 30, 11])
    for s in D['states']:
        _row(ws, r, [s['state_id'], s['state_type'], s['definition'], s['criteria'],
                     ' '.join(map(str, s['source_rows'])), s['used_tc_count']], mono_cols=(1,))
        r += 1

    # ── 4. Transition Dictionary ────────────────────────────────────────────
    ws, r = _sheet(wb, 'Transition Dictionary', '화면 간 이동 관계 정의서 (5.일반채팅)',
                   ['Transition ID', '시작 화면', 'Action', 'Target ID', '도착 화면', '설명', '출처 원본 행번호'],
                   [42, 26, 12, 26, 26, 34, 22])
    for t in D['transitions']:
        _row(ws, r, [t['transition_id'], t['start_screen'], t['action'], t['target_id'],
                     t['dest_screen'], t['description'], ' '.join(map(str, t['source_rows']))],
             mono_cols=(1, 2, 4, 5))
        r += 1

    # ── 5. 5.일반채팅 TC 리스트 (TC + Step) ─────────────────────────────────
    headers = ['New TC ID', 'TC명', '검증 목적', '우선순위', 'OS', '변환 유형', '자동화 준비',
               '원본 행번호', 'Step', 'start_screen',
               'precond·Data', 'precond·UI', 'precond·Permission',
               'target', 'target_parameter', 'action', 'action_parameter',
               'result_screen', 'result_state']
    widths = [15, 34, 40, 8, 8, 13, 14, 12, 22,
              20, 22, 22, 22, 26, 20, 12, 24, 20, 46]
    ws, r = _sheet(wb, '5.일반채팅 TC 리스트', '기존 TC → 신규 구조화 Step 변환 결과',
                   headers, widths)
    # TC 정렬: 최소 원본행 → new_tc_id (대시보드와 동일)
    tcs = sorted(D['tcs'], key=lambda t: (min(t['source_row_numbers']) if t['source_row_numbers'] else 0,
                                          t['new_tc_id']))
    gi = 0
    for t in tcs:
        fill = GROUP_FILL_A if gi % 2 == 0 else GROUP_FILL_B
        gi += 1
        rows_txt = ' '.join(map(str, t['source_row_numbers']))
        for s in t['steps']:
            vals = [t['new_tc_id'], t['title'], t['purpose'], t['priority'], t['os'],
                    t['conversion_type'], t['automation_readiness'], rows_txt,
                    s['step_no'], s['start_screen'],
                    s['precondition_data_state'], s['precondition_ui_state'],
                    s['precondition_permission_state'], s['target'], s['target_parameter'],
                    s['action'], s['action_parameter'], s['result_screen'], s['result_state']]
            _row(ws, r, vals, mono_cols=(1, 10, 14, 18))
            for c in range(1, len(headers) + 1):
                if ws.cell(r, c).fill.fgColor.rgb in ('00000000', None):
                    ws.cell(r, c).fill = fill
                else:
                    ws.cell(r, c).fill = fill
            r += 1

    # ── 6. 변환 요약 ────────────────────────────────────────────────────────
    ws, r = _sheet(wb, '변환 요약', '변환/검증 요약 (conversion_summary)',
                   ['항목', '값'], [40, 20])
    label = {
        'source_sheet': '대상 시트', 'source_tc_count': '원본 TC 수',
        'converted_tc_count': '신규 TC 수', 'converted_step_count': '신규 Step 수',
        'screen_dictionary_count': 'Screen 사전 수', 'target_dictionary_count': 'Target 사전 수',
        'state_dictionary_count': 'State 사전 수', 'transition_dictionary_count': 'Transition 사전 수',
        'mapped_source_tc_count': '연결 완료 원본 TC', 'unmapped_source_tc_count': '미연결 원본 TC',
        'split_source_tc_count': '분리(SPLIT)', 'merged_source_tc_count': '통합(MERGED)',
        'automation_ready_count': '자동화 READY', 'review_required_count': 'REVIEW_REQUIRED',
        'manual_only_count': 'MANUAL_ONLY', 'validation_passed': '검증 통과',
        'validation_total': '검증 항목 수', 'status': '상태',
    }
    for k, ko in label.items():
        _row(ws, r, [ko, D['summary'].get(k, '')])
        r += 1

    wb.save(OUT)
    print(f'저장 완료: {OUT}')
    print(f'  탭: Screen({len(D["screens"])}) / Target({len(D["targets"])}) / '
          f'State({len(D["states"])}) / Transition({len(D["transitions"])}) / '
          f'TC 리스트({len(D["tcs"])} TC, {D["summary"]["converted_step_count"]} Step)')


if __name__ == '__main__':
    build()
